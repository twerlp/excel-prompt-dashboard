#!/usr/bin/env python3
"""
Import a user story dataset (Excel) into Qdrant (vectors) and SQLite (metadata).

Expected Excel sheets and columns:
  "User Stories": Story ID, Title, Domain, Priority, Story Points,
                  Sprint, Status, Dependencies, User Story, Acceptance Criteria
  "Functional Requirements": Story ID, Title, FR Reference,
                             Functional Requirement Specification
  "Test Cases": Story ID, Title, Test Case ID, Test Description,
                Test Type, Test Steps, Expected Result

Usage:
  python import_dataset.py path/to/Dataset.xlsx
  python import_dataset.py path/to/Dataset.xlsx --mode upsert
  python import_dataset.py path/to/Dataset.xlsx --limit 5 --dry-run
  python import_dataset.py path/to/Dataset.xlsx --mapping column_map.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import yaml

from evaluation_kb import KnowledgeBase
from qdrant_store import QdrantStore


# ── Column name mapping ────────────────────────────────────────────────

DEFAULT_COLUMNS = {
    "user_stories": {
        "story_id": "Story ID",
        "title": "Title",
        "domain": "Domain",
        "priority": "Priority",
        "story_points": "Story Points",
        "sprint": "Sprint",
        "status": "Status",
        "dependencies": "Dependencies",
        "user_story": "User Story",
        "acceptance_criteria": "Acceptance Criteria",
    },
    "functional_requirements": {
        "story_id": "Story ID",
        "title": "Title",
        "fr_ref": "FR Reference",
        "fr_text": "Functional Requirement Specification",
    },
    "test_cases": {
        "story_id": "Story ID",
        "title": "Title",
        "tc_id": "Test Case ID",
        "description": "Test Description",
        "test_type": "Test Type",
        "steps": "Test Steps",
        "expected": "Expected Result",
    },
}

REQUIRED_COLUMNS = {
    "user_stories": ["story_id", "title", "user_story", "acceptance_criteria"],
    "functional_requirements": ["story_id", "fr_ref", "fr_text"],
    "test_cases": ["story_id", "tc_id", "description", "test_type"],
}


# ── Config ─────────────────────────────────────────────────────────────

def load_config() -> Dict[str, Any]:
    config_path = Path(__file__).resolve().parent / "config.yaml"
    if config_path.exists():
        with open(config_path) as f:
            return yaml.safe_load(f) or {}
    return {}


# ── Parsing ────────────────────────────────────────────────────────────

def _get_column_map(sheet_config: Dict[str, str],
                    header_row: List[Optional[str]]) -> Dict[str, int]:
    """Map internal keys to column indices from the header row."""
    col_idx: Dict[str, int] = {}
    clean_headers = [str(h).strip() if h else "" for h in header_row]
    for key, col_name in sheet_config.items():
        for i, h in enumerate(clean_headers):
            if h.lower() == col_name.lower():
                col_idx[key] = i
                break
    return col_idx


def parse_user_stories(ws, col_map: Dict[str, str],
                       validation_rules: Dict[str, Any]) -> Tuple[List[Dict], List[str]]:
    """Parse the User Stories sheet. Returns (stories, warnings)."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = _get_column_map(col_map, headers)

    missing = [k for k in REQUIRED_COLUMNS["user_stories"] if k not in col_idx]
    if missing:
        return [], [f"Missing required columns in User Stories sheet: {missing}"]

    stories = []
    warnings = []
    valid_priorities = validation_rules.get("valid_priorities", [])
    valid_statuses = validation_rules.get("valid_statuses", [])

    for row in range(2, ws.max_row + 1):
        def val(key):
            c = col_idx.get(key)
            if c is not None:
                return ws.cell(row=row, column=c + 1).value
            return None

        sid = val("story_id")
        title = val("title")
        if not sid or not title:
            continue

        sid = str(sid).strip()
        story = {
            "story_id": sid,
            "title": str(title).strip(),
            "domain": str(val("domain") or "").strip(),
            "priority": str(val("priority") or "").strip(),
            "story_points": int(val("story_points") or 0),
            "sprint": str(val("sprint") or "").strip(),
            "status": str(val("status") or "").strip(),
            "dependencies": str(val("dependencies") or "").strip(),
            "user_story": str(val("user_story") or "").strip(),
            "acceptance_criteria": str(val("acceptance_criteria") or "").strip(),
        }
        story["story_text"] = story["user_story"] + "\n" + story["acceptance_criteria"]

        if story["priority"] and valid_priorities and story["priority"] not in valid_priorities:
            warnings.append(f"Story {sid}: unknown priority '{story['priority']}'")
        if story["status"] and valid_statuses and story["status"] not in valid_statuses:
            warnings.append(f"Story {sid}: unknown status '{story['status']}'")

        stories.append(story)

    return stories, warnings


def parse_functional_requirements(ws, col_map: Dict[str, str],
                                   fr_pattern: str) -> Tuple[List[Dict], List[str]]:
    """Parse the Functional Requirements sheet. Returns (frs, warnings)."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = _get_column_map(col_map, headers)

    missing = [k for k in REQUIRED_COLUMNS["functional_requirements"] if k not in col_idx]
    if missing:
        return [], [f"Missing required columns in Functional Requirements sheet: {missing}"]

    frs = []
    warnings = []

    for row in range(2, ws.max_row + 1):
        def val(key):
            c = col_idx.get(key)
            if c is not None:
                return ws.cell(row=row, column=c + 1).value
            return None

        sid = val("story_id")
        fr_ref = val("fr_ref")
        fr_text = val("fr_text")
        if not sid or not fr_text:
            continue

        fr_ref_clean = str(fr_ref).strip() if fr_ref else ""
        if fr_pattern and fr_ref_clean and not re.match(fr_pattern, fr_ref_clean):
            warnings.append(f"FR '{fr_ref_clean}': does not match pattern {fr_pattern}")

        frs.append({
            "story_id": str(sid).strip(),
            "fr_ref": fr_ref_clean,
            "fr_text": str(fr_text).strip(),
        })

    return frs, warnings


def parse_test_cases(ws, col_map: Dict[str, str],
                      tc_pattern: str) -> Tuple[List[Dict], List[str]]:
    """Parse the Test Cases sheet. Returns (test_cases, warnings)."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = _get_column_map(col_map, headers)

    missing = [k for k in REQUIRED_COLUMNS["test_cases"] if k not in col_idx]
    if missing:
        return [], [f"Missing required columns in Test Cases sheet: {missing}"]

    tcs = []
    warnings = []

    for row in range(2, ws.max_row + 1):
        def val(key):
            c = col_idx.get(key)
            if c is not None:
                return ws.cell(row=row, column=c + 1).value
            return None

        sid = val("story_id")
        tc_id = val("tc_id")
        description = val("description")
        if not sid or not description:
            continue

        tc_id_clean = str(tc_id).strip() if tc_id else ""
        if tc_pattern and tc_id_clean and not re.match(tc_pattern, tc_id_clean):
            warnings.append(f"TC '{tc_id_clean}': does not match pattern {tc_pattern}")

        tcs.append({
            "story_id": str(sid).strip(),
            "tc_id": tc_id_clean,
            "description": str(description).strip(),
            "test_type": str(val("test_type") or "").strip(),
            "steps": str(val("steps") or "").strip(),
            "expected": str(val("expected") or "").strip(),
        })

    return tcs, warnings


def validate_cross_references(stories: List[Dict], frs: List[Dict],
                               tcs: List[Dict]) -> List[str]:
    """Check that all FR and TC story_ids exist in the stories list."""
    errors = []
    story_ids = {s["story_id"] for s in stories}
    for fr in frs:
        if fr["story_id"] not in story_ids:
            errors.append(f"FR for story '{fr['story_id']}' not found in User Stories")
    for tc in tcs:
        if tc["story_id"] not in story_ids:
            errors.append(f"TC for story '{tc['story_id']}' not found in User Stories")
    return errors


# ── Import ─────────────────────────────────────────────────────────────

def import_dataset(excel_path: str, mode: str = "skip",
                   limit: Optional[int] = None, dry_run: bool = False,
                   column_mapping: Optional[str] = None) -> Dict[str, int]:
    """
    Main import function. Parses Excel, validates, stores in both DBs.

    Returns: {stories, fr_count, tc_count, warnings, errors}
    """
    config = load_config()
    import_cfg = config.get("import", {})
    validation_rules = {
        "fr_ref_pattern": import_cfg.get("fr_ref_pattern", r"^FR-\d{3}$"),
        "tc_id_pattern": import_cfg.get("tc_id_pattern", r"^TC-\d{3}-\d{2}$"),
        "valid_priorities": import_cfg.get("valid_priorities", []),
        "valid_statuses": import_cfg.get("valid_statuses", []),
    }
    qdrant_cfg = config.get("qdrant", {})
    db_cfg = config.get("database", {})

    # Column mapping
    col_maps = DEFAULT_COLUMNS.copy()
    if column_mapping:
        with open(column_mapping) as f:
            custom = json.load(f)
        for sheet, cols in custom.items():
            if sheet in col_maps:
                col_maps[sheet].update(cols)

    # Open workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = wb.sheetnames

    # Find sheets (flexible naming)
    us_sheet = next((s for s in sheet_names if "user stori" in s.lower()), sheet_names[0])
    fr_sheet = next((s for s in sheet_names if "functional" in s.lower()), None)
    tc_sheet = next((s for s in sheet_names if "test case" in s.lower()), None)

    all_warnings: List[str] = []
    all_errors: List[str] = []

    # Parse stories
    stories, w = parse_user_stories(
        wb[us_sheet], col_maps["user_stories"], validation_rules
    )
    all_warnings.extend(w)
    if not stories:
        all_errors.append("No user stories found")

    if limit:
        stories = stories[:limit]

    # Parse FRs
    frs: List[Dict] = []
    if fr_sheet:
        frs, w = parse_functional_requirements(
            wb[fr_sheet], col_maps["functional_requirements"],
            validation_rules["fr_ref_pattern"]
        )
        all_warnings.extend(w)

    # Parse TCs
    tcs: List[Dict] = []
    if tc_sheet:
        tcs, w = parse_test_cases(
            wb[tc_sheet], col_maps["test_cases"],
            validation_rules["tc_id_pattern"]
        )
        all_warnings.extend(w)

    # Cross-reference check
    all_errors.extend(validate_cross_references(stories, frs, tcs))

    if all_errors:
        print(f"ERRORS ({len(all_errors)}):")
        for e in all_errors:
            print(f"  - {e}")
        if not dry_run:
            return {
                "stories": 0, "fr_count": 0, "tc_count": 0,
                "warnings": len(all_warnings), "errors": len(all_errors),
            }

    if dry_run:
        print(f"\nDry run — would import:")
        print(f"  Stories: {len(stories)}")
        print(f"  Functional Requirements: {len(frs)}")
        print(f"  Test Cases: {len(tcs)}")
        if all_warnings:
            print(f"  Warnings: {len(all_warnings)}")
            for w in all_warnings[:10]:
                print(f"    - {w}")
        return {
            "stories": len(stories), "fr_count": len(frs), "tc_count": len(tcs),
            "warnings": len(all_warnings), "errors": 0,
        }

    # Store in databases
    kb = KnowledgeBase(db_cfg.get("sqlite_path", "kb.sqlite"))
    kb.initialize()

    qdrant_path = qdrant_cfg.get("path", "./qdrant_data")
    store = QdrantStore(path=qdrant_path)
    store.ensure_collections()

    stories_inserted = 0
    for story in stories:
        if kb.insert_story(story, mode=mode):
            stories_inserted += 1

    fr_inserted = kb.insert_frs(frs, source="ground_truth")
    tc_inserted = kb.insert_test_cases(tcs, source="ground_truth")

    store.upsert_stories(stories)

    kb.record_import(
        filename=Path(excel_path).name,
        stories_count=stories_inserted,
        fr_count=fr_inserted,
        tc_count=tc_inserted,
        warnings=len(all_warnings),
        errors=len(all_errors),
    )

    kb.close()
    wb.close()

    return {
        "stories": stories_inserted,
        "fr_count": fr_inserted,
        "tc_count": tc_inserted,
        "warnings": len(all_warnings),
        "errors": len(all_errors),
    }


# ── CLI ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Import user story Excel dataset into Qdrant + SQLite"
    )
    parser.add_argument("excel_path", help="Path to the Excel file")
    parser.add_argument("--mode", default="skip", choices=["skip", "upsert"],
                        help="Skip duplicates or upsert (default: skip)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limit number of stories to import")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate only, do not write to databases")
    parser.add_argument("--mapping", default=None,
                        help="JSON file mapping expected column names to actual names")
    args = parser.parse_args()

    path = Path(args.excel_path)
    if not path.exists():
        print(f"Error: file not found: {args.excel_path}")
        sys.exit(1)

    result = import_dataset(
        excel_path=str(path),
        mode=args.mode,
        limit=args.limit,
        dry_run=args.dry_run,
        column_mapping=args.mapping,
    )

    print(f"\nImport summary:")
    print(f"  Stories:     {result['stories']}")
    print(f"  FRs:         {result['fr_count']}")
    print(f"  Test Cases:  {result['tc_count']}")
    print(f"  Warnings:    {result['warnings']}")
    print(f"  Errors:      {result['errors']}")


if __name__ == "__main__":
    main()
