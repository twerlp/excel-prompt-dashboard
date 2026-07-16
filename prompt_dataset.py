#!/usr/bin/env python3
"""
Automated FRS & Test Case Generation via LLM Prompting
======================================================
Reads user stories from `generate_dataset.py`, applies configurable prompt
templates via pluggable LLM backends, and writes results back to the Excel
dataset with full metadata (model, provider, prompt hash, timestamp).

Extensibility:
  - Add new prompt templates in PROMPT_TEMPLATES.
  - Register new LLM backends via the `LlmBackend` abstract base class.
   - Insert new user stories via the `--stories` CLI flag (reads a JSON file).

Usage:
  python prompt_dataset.py                # run with default (internal) backend
  OPENAI_API_KEY=sk-... python prompt_dataset.py --backend openai
  python prompt_dataset.py --output my_dataset.xlsx --stories new_stories.json
"""

from __future__ import annotations

import hashlib
import json
import os
import sys
from abc import ABC, abstractmethod
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
# 1. Prompt Templates
# ═══════════════════════════════════════════════════════════════════════════

PROMPT_TEMPLATES: Dict[str, Dict[str, str]] = {
    # ── FRS PROMPTS ──────────────────────────────────────────────────────
    "frs_zero_shot": {
        "category": "frs",
        "name": "Zero-Shot Structured FRS",
        "description": "Direct instruction to produce FRS in the canonical FR-XXX format, "
                       "using 'The system shall…' language with a sequential numbering scheme.",
        "system_prompt": (
            "You are a senior systems analyst and requirements engineer. "
            "Write functional requirement specifications (FRS) that are precise, "
            "testable, and implementation-agnostic where possible."
        ),
        "user_prompt_template": (
            "Given the user story and its acceptance criteria below, produce a complete "
            "Functional Requirement Specification.\n\n"
            "## Rules\n"
            "1. Number each requirement sequentially starting from FR-{next_fr_num}.\n"
            "2. Use the canonical form: \"FR-XXX: The system shall …\"\n"
            "3. Each acceptance criterion must be traceable to at least one FR.\n"
            "4. Be specific about data formats, validation rules, error states, and constraints.\n"
            "5. Include requirements for security, audit, and error handling where applicable.\n"
            "6. Prefix every requirement line with \"FR-\" and make each a complete sentence.\n\n"
            "## User Story\n"
            "**Title:** {title}\n"
            "**Domain:** {domain}\n"
            "**Priority:** {priority}\n"
            "**User Story:** {user_story}\n\n"
            "## Acceptance Criteria\n"
            "{acceptance_criteria}\n\n"
            "## Dependencies (context only — do not write requirements for these)\n"
            "{dependencies}\n\n"
            "Output ONLY the FRS lines, one per line, with no additional commentary."
        ),
    },
    "frs_few_shot": {
        "category": "frs",
        "name": "Few-Shot with Examples",
        "description": "Provide 3 curated example FRS blocks from different domains "
                       "so the model learns the expected depth, tone, and formatting.",
        "system_prompt": (
            "You are a principal software architect writing functional requirements for "
            "a development team. Follow the exact format and level of detail shown in "
            "the examples."
        ),
        "user_prompt_template": (
            "## Example 1 (E-Commerce — Shopping Cart)\n"
            "FR-009: The system shall maintain a cart data structure keyed by anonymous session ID "
            "for guest users and by user ID for authenticated users.\n"
            "FR-010: The system shall merge a guest cart into the user's persistent cart upon login.\n"
            "FR-011: The system shall provide REST endpoints: POST /cart/items, "
            "PATCH /cart/items/{{sku}}, DELETE /cart/items/{{sku}}.\n\n"
            "## Example 2 (Banking — Funds Transfer)\n"
            "FR-036: The system shall provide a /transfer endpoint accepting source_account_id, "
            "destination_account_id, amount, currency, and note.\n"
            "FR-037: The system shall execute intra-bank transfers synchronously within a database "
            "transaction — debit source, credit destination, insert transaction record.\n\n"
            "## Example 3 (Healthcare — Appointment Booking)\n"
            "FR-043: The system shall expose GET /doctors/{{id}}/slots returning available time windows "
            "with status (available/locked/booked).\n"
            "FR-044: When a patient selects a slot, the system shall set a pessimistic lock "
            "(status=locked, locked_until=now+5min) to prevent double-booking.\n\n"
            "## Task\n"
            "Write the FRS for the user story below. Start numbering at FR-{next_fr_num}.\n\n"
            "**Title:** {title}  |  **Domain:** {domain}  |  **Priority:** {priority}\n"
            "**User Story:** {user_story}\n\n"
            "**Acceptance Criteria:**\n{acceptance_criteria}\n\n"
            "**Dependencies:** {dependencies}\n\n"
            "Output ONLY the FR lines. Do NOT include headings or commentary."
        ),
    },
    "frs_cot": {
        "category": "frs",
        "name": "Chain-of-Thought FRS",
        "description": "Ask the model to reason through the acceptance criteria step by step "
                       "before writing the final FRs, improving completeness.",
        "system_prompt": (
            "You are an expert business analyst. Think through each acceptance criterion "
            "carefully before writing the corresponding functional requirements."
        ),
        "user_prompt_template": (
            "For the user story below, produce a Functional Requirement Specification.\n\n"
            "**User Story:** {user_story}\n\n"
            "**Acceptance Criteria:**\n{acceptance_criteria}\n\n"
            "**Domain/Priority:** {domain} / {priority}\n"
            "**Dependencies:** {dependencies}\n\n"
            "## Step 1 — Analyse each acceptance criterion. "
            "What systems, APIs, validations, data models, and error states are implied?\n\n"
            "## Step 2 — For each implied need, write a functional requirement numbered "
            "FR-{next_fr_num} onwards. Use \"The system shall…\" language. "
            "Be specific about HTTP methods, field types, constraints, and error messages.\n\n"
            "## Step 3 — Review: does every acceptance criterion map to at least one FR? "
            "Are edge cases (timeouts, concurrent access, invalid input, unauthorized access) covered?\n\n"
            "Output ONLY the final FR lines (one per line). No analysis or step labels."
        ),
    },
    # ── TEST-CASE PROMPTS ─────────────────────────────────────────────────
    "tc_zero_shot": {
        "category": "test_case",
        "name": "Zero-Shot Structured Test Cases",
        "description": "Direct instruction to produce test cases covering happy path, "
                       "error states, boundary values, and security concerns.",
        "system_prompt": (
            "You are a senior QA engineer writing manual test cases. "
            "Follow the exact JSON structure provided and cover functional, boundary, "
            "security, and integration scenarios."
        ),
        "user_prompt_template": (
            "Generate test cases for the following user story.\n\n"
            "**Title:** {title}  |  **Domain:** {domain}  |  **Priority:** {priority}\n"
            "**User Story:** {user_story}\n\n"
            "**Acceptance Criteria:**\n{acceptance_criteria}\n\n"
            "**Functional Requirements (context):**\n{frs}\n\n"
            "## Rules\n"
            "1. Number test cases TC-{next_tc_prefix}-01, TC-{next_tc_prefix}-02, …\n"
            "2. Each test case must have: tc_id, description, test_type, steps, expected.\n"
            "3. Test types must include 'Functional' + at least one of 'Boundary', "
            "'Security', 'Performance', 'Integration', or 'Reliability'.\n"
            "4. Steps are numbered, actionable, and reference specific UI elements or API calls.\n"
            "5. Expected results are specific and measurable.\n"
            "6. Cover at minimum: (a) happy path, (b) each acceptance criterion, "
            "(c) at least 2 error/negative paths, (d) at least 1 security concern.\n"
            "7. Produce 4–6 test cases total.\n\n"
            "Output ONLY a JSON array of test case objects with keys: "
            "tc_id, description, test_type, steps, expected."
        ),
    },
    "tc_few_shot": {
        "category": "test_case",
        "name": "Few-Shot with Examples",
        "description": "Provide example test cases from different domains to illustrate "
                       "the expected format, depth, and variety of test types.",
        "system_prompt": (
            "You are a QA lead. Write detailed manual test cases matching the format "
            "and quality level shown in the examples."
        ),
        "user_prompt_template": (
            "## Example TC (from a Registration story)\n"
            '{{\n'
            '  "tc_id": "TC-001-02",\n'
            '  "description": "Verify duplicate email is rejected.",\n'
            '  "test_type": "Functional",\n'
            '  "steps": "1. Register with email already in use.\\n2. Submit form.",\n'
            '  "expected": "Error: \\"An account with this email already exists.\\" No second record."\n'
            '}}\n'
            "## Example TC (from a Payment story)\n"
            '{{\n'
            '  "tc_id": "TC-003-03",\n'
            '  "description": "Verify idempotency — double-click protection.",\n'
            '  "test_type": "Functional",\n'
            '  "steps": "1. Submit payment.\\n2. Resend identical request before first completes.",\n'
            '  "expected": "Only one PaymentIntent created. Both responses return the same outcome."\n'
            '}}\n\n'
            "## Task\n"
            "Generate test cases for:\n"
            "**Story:** {title} ({domain}, {priority})\n"
            "**User Story:** {user_story}\n"
            "**Acceptance Criteria:**\n{acceptance_criteria}\n"
            "**FRS:**\n{frs}\n\n"
            "Start IDs at TC-{next_tc_prefix}-01. Output ONLY a JSON array."
        ),
    },
    "tc_cot": {
        "category": "test_case",
        "name": "Chain-of-Thought Test Cases",
        "description": "Reason through each acceptance criterion and FR to derive "
                       "test scenarios before writing final test cases.",
        "system_prompt": (
            "You are a test architect. First think through testability of each "
            "acceptance criterion, then write precise, repeatable test cases."
        ),
        "user_prompt_template": (
            "## Task\n"
            "Generate test cases for the user story below.\n\n"
            "**User Story:** {user_story}\n"
            "**Acceptance Criteria:**\n{acceptance_criteria}\n"
            "**FRS:**\n{frs}\n\n"
            "## Analysis (do NOT include in output)\n"
            "- For each acceptance criterion, identify the positive test, the negative "
            "test, and any edge/boundary condition.\n"
            "- Check if any FRS implies security, performance, or integration tests.\n"
            "- Ensure full coverage: every criterion and FR has at least one test.\n\n"
            "## Output\n"
            "Produce 4–6 test cases numbered TC-{next_tc_prefix}-01 onwards. "
            "Each as a JSON object with keys: tc_id, description, test_type, steps, expected.\n"
            "Cover these scenarios: happy path, validation error, boundary value, "
            "authorization check, and one reliability/integration test.\n\n"
            "Output ONLY the JSON array."
        ),
    },
}


# ═══════════════════════════════════════════════════════════════════════════
# 2. LLM Backend Abstraction
# ═══════════════════════════════════════════════════════════════════════════

class LlmBackend(ABC):
    """Abstract interface for any LLM service provider."""

    @abstractmethod
    def complete(self, system_prompt: str, user_prompt: str,
                 metadata: Optional[Dict[str, str]] = None) -> str:
        """Return the completion text for the given prompts.

        Args:
            system_prompt: The system-level instruction for the LLM.
            user_prompt: The user-facing prompt with story details.
            metadata: Optional dict with keys 'story_id' and 'category'
                      (e.g., 'frs_few_shot', 'tc_few_shot') for backends
                      that route or look up pre-generated content.
        """
        ...

    @property
    @abstractmethod
    def model_name(self) -> str:
        ...

    @property
    @abstractmethod
    def provider_name(self) -> str:
        ...


class InternalLlmBackend(LlmBackend):
    """
    Generates content programmatically using structured templates.
    Used when no external API is available.  Produces deterministic,
    well-structured FRS and test cases from parsed acceptance criteria.

    Note: This backend is designed for demo and fallback scenarios.
    For production-quality output, use a live LLM backend (OpenAI, Anthropic).
    """

    # Domain-specific FR patterns for richer output
    _DOMAIN_VERBS = {
        "E-Commerce": ["provide", "display", "validate", "calculate"],
        "Banking": ["execute", "authorize", "validate", "log"],
        "Healthcare": ["schedule", "protect", "notify", "generate"],
        "Cybersecurity": ["authenticate", "encrypt", "log", "enforce"],
        "DevOps": ["trigger", "deploy", "report", "block"],
        "Finance": ["generate", "compute", "distribute", "record"],
        "Social Media": ["rank", "filter", "cache", "load"],
        "Enterprise SaaS": ["enforce", "assign", "intercept", "log"],
        "Analytics / BI": ["stream", "aggregate", "display", "cache"],
        "HR & Payroll": ["validate", "route", "notify", "deduct"],
        "Supply Chain": ["upload", "parse", "validate", "upsert"],
        "Mobile / Engagement": ["sync", "filter", "store", "deliver"],
        "Platform / Infrastructure": ["throttle", "enforce", "return", "log"],
        "Data Privacy / Compliance": ["compile", "encrypt", "confirm", "redact"],
        "UI/UX": ["apply", "persist", "transition", "detect"],
        "Customer Support": ["relay", "create", "lookup", "queue"],
    }

    _DEFAULT_VERBS = ["provide", "support", "validate", "process", "handle", "implement"]

    def __init__(self, model: str = "internal-template-v2"):
        self._model = model

    @property
    def model_name(self) -> str:
        return self._model

    @property
    def provider_name(self) -> str:
        return "RuleBased"

    def complete(self, system_prompt: str, user_prompt: str,
                 metadata: Optional[Dict[str, str]] = None) -> str:
        import re

        category = (metadata or {}).get("category", "")
        if category.startswith("frs") or "Functional Requirement Specification" in user_prompt:
            return self._generate_frs(user_prompt)
        elif category.startswith("tc") or "test case" in system_prompt.lower():
            return self._generate_test_cases(user_prompt)
        return ""

    def _parse_story_context(self, prompt: str) -> Dict[str, str]:
        """Extract story metadata from the formatted user prompt."""
        import re
        ctx: Dict[str, str] = {}
        for field in ["Title", "Domain", "Priority", "User Story"]:
            m = re.search(rf'\*\*{field}[:\*]*\*\*\s*(.*?)(?:\n|$)', prompt)
            if m:
                ctx[field.lower().replace(" ", "_")] = m.group(1).strip()
        ac_section = ""
        if "Acceptance Criteria" in prompt:
            parts = prompt.split("Acceptance Criteria")
            if len(parts) > 1:
                ac_section = parts[1].split("##")[0].split("**Dependencies")[0]
        ctx["acceptance_criteria"] = ac_section
        return ctx

    def _extract_ac_items(self, prompt: str) -> List[str]:
        import re
        ac_section = ""
        if "Acceptance Criteria" in prompt:
            parts = prompt.split("Acceptance Criteria")
            if len(parts) > 1:
                ac_section = parts[1].split("##")[0].split("**Dependencies")[0]
        return [a.strip() for a in re.findall(r'\d+\.\s*(.*?)(?=\d+\.|\Z)', ac_section, re.DOTALL) if a.strip()]

    def _extract_key_entities(self, text: str) -> List[str]:
        """Pull out key nouns/verbs from text for richer FR wording."""
        import re
        entities = []
        # Look for API paths, UI elements, data types
        for pattern in [
            r'(?:POST|GET|PUT|DELETE|PATCH)\s+(/\S+)',
            r'(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?\s*(?:KB|MB|GB|hours?|minutes?|days?|seconds?|ms))',
            r'([A-Z][a-z]+(?:\s[A-Z][a-z]+){2,})',
            r'(?:redirect|navigate)\s+(?:to|from)\s+(\S+)',
        ]:
            for match in re.finditer(pattern, text):
                val = match.group(1).strip().rstrip(".,;:!?")
                if val and len(val) > 2:
                    entities.append(val)
        return entities[:6]  # cap for brevity

    def _generate_frs(self, prompt: str) -> str:
        import re
        ctx = self._parse_story_context(prompt)
        ac_items = self._extract_ac_items(prompt)
        domain = ctx.get("domain", "")
        user_story = ctx.get("user_story", "")
        verbs = self._DOMAIN_VERBS.get(domain, self._DEFAULT_VERBS)
        entities = self._extract_key_entities(prompt)

        next_fr = 200
        m = re.search(r'FR-(\d+)', prompt)
        if m:
            next_fr = int(m.group(1))

        lines: List[str] = []
        fr_idx = next_fr

        # Generate FRs based on acceptance criteria plus inferred system concerns
        for i, ac in enumerate(ac_items):
            verb = verbs[i % len(verbs)]
            # Clean up the AC text into something more FR-like
            ac_clean = ac.rstrip(".").strip()
            # Try to make it more specific by incorporating entities
            entity_hint = ""
            if entities and i < len(entities):
                entity_hint = f" ({entities[i]})"

            # Translate common AC patterns into FR language
            ac_lower = ac_clean.lower()
            if "valid" in ac_lower or "format" in ac_lower or "must " in ac_lower:
                line = (f"FR-{fr_idx:03d}: The system shall validate and enforce "
                        f"that {ac_clean[0].lower()}{ac_clean[1:]}.")
            elif "email" in ac_lower or "notification" in ac_lower or "confirm" in ac_lower:
                line = (f"FR-{fr_idx:03d}: The system shall {verb} {ac_clean[0].lower()}{ac_clean[1:]}"
                        f"{entity_hint} and log the dispatch event.")
            elif "send" in ac_lower or "sent" in ac_lower or "return" in ac_lower:
                line = (f"FR-{fr_idx:03d}: The system shall {verb} {ac_clean[0].lower()}{ac_clean[1:]}"
                        f" within SLA timelines and handle delivery failures with retry logic{entity_hint}.")
            elif "prevent" in ac_lower or "block" in ac_lower or "deny" in ac_lower:
                line = (f"FR-{fr_idx:03d}: The system shall enforce that {ac_clean[0].lower()}{ac_clean[1:]}"
                        f" and return an appropriate error response{entity_hint}.")
            else:
                line = (f"FR-{fr_idx:03d}: The system shall {verb} functionality to "
                        f"{ac_clean[0].lower()}{ac_clean[1:]}{entity_hint}.")
            lines.append(line)
            fr_idx += 1

        # Add a catch-all FR for edge cases and error handling
        lines.append(
            f"FR-{fr_idx:03d}: The system shall handle error conditions gracefully "
            f"for all operations described above, returning structured error responses "
            f"with HTTP-appropriate status codes and logging diagnostic information."
        )

        if not lines:
            lines.append(f"FR-{next_fr:03d}: The system shall fulfill all specified acceptance criteria.")
        return "\n".join(lines)

    def _generate_test_cases(self, prompt: str) -> str:
        import re
        tc_prefix = "GEN"
        m = re.search(r'TC-(\S+)-', prompt)
        if m:
            tc_prefix = m.group(1)

        ac_items = self._extract_ac_items(prompt)
        tcs = []

        # Determine which test types to use based on the acceptance criteria
        types_pool = ["Functional", "Functional", "Boundary", "Security", "Integration", "Performance"]
        if len(ac_items) > 4:
            types_pool = ["Functional", "Functional", "Boundary", "Security", "Integration", "Performance"]

        for i, ac in enumerate(ac_items[:6]):
            ac_short = ac[:80].rstrip(".").strip()
            test_type = types_pool[i] if i < len(types_pool) else "Functional"

            # Build more specific steps and expected results
            steps = self._build_test_steps(ac, i, test_type)
            expected = self._build_test_expected(ac, test_type)

            tc = {
                "tc_id": f"TC-{tc_prefix}-{i+1:02d}",
                "description": f"Verify that: {ac_short}",
                "test_type": test_type,
                "steps": steps,
                "expected": expected,
            }
            tcs.append(tc)

        return json.dumps(tcs, indent=2)

    def _build_test_steps(self, ac_text: str, index: int, test_type: str) -> str:
        """Build numbered test steps from an acceptance criterion."""
        ac_lower = ac_text.lower()
        steps_parts = []

        if test_type == "Security":
            steps_parts = [
                f"1. Authenticate as a user with appropriate permissions.",
                f"2. Attempt the operation described in: {ac_text[:60]}",
                f"3. Observe the system response, including HTTP status and body.",
                f"4. Inspect audit logs for the security event.",
            ]
        elif test_type == "Boundary":
            steps_parts = [
                f"1. Identify the boundary condition implied by: {ac_text[:60]}",
                f"2. Set up the system state at or beyond the boundary threshold.",
                f"3. Execute the boundary-triggering action.",
                f"4. Observe the system behaviour and error handling.",
            ]
        elif test_type == "Performance":
            steps_parts = [
                f"1. Warm up the system with a baseline load.",
                f"2. Execute the operation under expected load conditions.",
                f"3. Measure response time, throughput, and error rate.",
                f"4. Compare metrics against SLA thresholds.",
            ]
        elif test_type == "Integration":
            steps_parts = [
                f"1. Configure all dependent services in test mode.",
                f"2. Trigger the operation that spans multiple services.",
                f"3. Verify each service receives and processes the expected data.",
                f"4. Simulate a downstream service failure and verify graceful degradation.",
            ]
        else:  # Functional
            if "log" in ac_lower or "login" in ac_lower or "sign" in ac_lower:
                steps_parts = [
                    f"1. Navigate to the relevant page or endpoint.",
                    f"2. Enter valid/test credentials or parameters.",
                    f"3. Submit the form or request.",
                    f"4. Verify the UI state or API response.",
                ]
            elif "error" in ac_lower or "invalid" in ac_lower or "reject" in ac_lower:
                steps_parts = [
                    f"1. Prepare invalid or edge-case input data.",
                    f"2. Submit the input via the appropriate interface.",
                    f"3. Observe the error response or validation message.",
                    f"4. Verify no side effects (database, state changes).",
                ]
            else:
                steps_parts = [
                    f"1. Set up the preconditions for criterion: {ac_text[:50]}",
                    f"2. Execute the primary scenario described in the acceptance criteria.",
                    f"3. Observe the system behaviour and output.",
                    f"4. Verify downstream effects (state changes, notifications, logs).",
                ]

        return "\n".join(steps_parts)

    def _build_test_expected(self, ac_text: str, test_type: str) -> str:
        """Build expected result from an acceptance criterion and test type."""
        ac_short = ac_text[:80].rstrip(".").strip()

        if test_type == "Security":
            return (f"The system blocks the unauthorized operation with HTTP 401/403. "
                    f"Audit log entry created with user ID, timestamp, IP, and action attempted. "
                    f"No data is leaked. Operation does not partially execute.")
        elif test_type == "Boundary":
            return (f"The system correctly handles the boundary condition specified in: "
                    f"'{ac_short}'. Either the operation is gracefully rejected with a "
                    f"descriptive message, or it is capped at the maximum allowed value.")
        elif test_type == "Performance":
            return (f"The system meets the performance SLA: response time within acceptable "
                    f"thresholds, error rate below 1%, and no resource exhaustion. "
                    f"Results are logged to the performance test report.")
        elif test_type == "Integration":
            return (f"All services in the integration chain process the request correctly. "
                    f"Data flows end-to-end without corruption. On downstream failure, "
                    f"the system returns an appropriate error and does not leave partial state.")
        else:  # Functional
            return (f"The system behaves according to the acceptance criterion: "
                    f"'{ac_short}'. The correct state transition occurs, "
                    f"appropriate feedback is shown, and no errors are logged.")




class PreGeneratedBackend(LlmBackend):
    """
    Returns pre-generated, high-quality LLM outputs stored in
    `pre_generated_llm_outputs.py`.  Used for demos and as a quality
    baseline for evaluating live LLM prompts.
    """
    def __init__(self, model: str = "pre-generated-v1"):
        self._model = model
        self._call_count = 0  # track FRS vs TC by alternation
        try:
            from pre_generated_llm_outputs import PRE_GENERATED  # type: ignore[import-untyped]
            self._db: Dict = PRE_GENERATED
        except ImportError:
            self._db = {}

    @property
    def model_name(self) -> str:
        return self._model

    @property
    def provider_name(self) -> str:
        return "PreGenerated"

    def complete(self, system_prompt: str, user_prompt: str,
                 metadata: Optional[Dict[str, str]] = None) -> str:
        story_id = (metadata or {}).get("story_id", "")
        category = (metadata or {}).get("category", "")
        key = (story_id, category)
        result = self._db.get(key)
        if result:
            return result
        alt_cat = "tc_few_shot" if category.startswith("frs") else "frs_few_shot"
        if not category.startswith(("frs", "tc")):
            alt_cat = "frs_few_shot"  # default fallback
        alt_result = self._db.get((story_id, alt_cat))
        if alt_result:
            return alt_result
        return f"// No pre-generated output for {story_id}/{category}. Use a live LLM backend.\n"


class OpenAiBackend(LlmBackend):
    """Backend for OpenAI-compatible APIs (GPT-4, GPT-3.5, etc.)."""

    def __init__(self, model: str = "gpt-4o", api_key: Optional[str] = None):
        self._model = model
        self._api_key = api_key or os.environ.get("OPENAI_API_KEY", "")

    @property
    def model_name(self) -> str:
        return self._model

    @property
    def provider_name(self) -> str:
        return "OpenAI"

    def complete(self, system_prompt: str, user_prompt: str,
                 metadata: Optional[Dict[str, str]] = None) -> str:
        try:
            from openai import OpenAI
        except ImportError:
            raise RuntimeError("Install openai: pip install openai")

        client = OpenAI(api_key=self._api_key)
        response = client.chat.completions.create(
            model=self._model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.2,
            max_tokens=4096,
        )
        return response.choices[0].message.content or ""


class AnthropicBackend(LlmBackend):
    """Backend for Anthropic's Claude models."""

    def __init__(self, model: str = "claude-sonnet-4-20250514", api_key: Optional[str] = None):
        self._model = model
        self._api_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")

    @property
    def model_name(self) -> str:
        return self._model

    @property
    def provider_name(self) -> str:
        return "Anthropic"

    def complete(self, system_prompt: str, user_prompt: str,
                 metadata: Optional[Dict[str, str]] = None) -> str:
        try:
            import anthropic
        except ImportError:
            raise RuntimeError("Install anthropic: pip install anthropic")

        client = anthropic.Anthropic(api_key=self._api_key)
        message = client.messages.create(
            model=self._model,
            max_tokens=4096,
            temperature=0.2,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )
        return message.content[0].text if message.content else ""


# Backend registry
BACKEND_REGISTRY: Dict[str, type] = {
    "pregenerated": PreGeneratedBackend,
    "internal": InternalLlmBackend,
    "openai": OpenAiBackend,
    "anthropic": AnthropicBackend,
}


# ═══════════════════════════════════════════════════════════════════════════
# 3. User Story Loader
# ═══════════════════════════════════════════════════════════════════════════

def load_user_stories() -> List[Dict[str, Any]]:
    """
    Load user stories from the bundled dataset in `generate_dataset.py`.
    In production, this could also read from an external JSON file.
    """
    # Import the stories list from the sibling script
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from generate_dataset import stories  # type: ignore[import-untyped]
    return stories


# ═══════════════════════════════════════════════════════════════════════════
# 4. Prompt Application
# ═══════════════════════════════════════════════════════════════════════════

def build_prompt(template: Dict[str, str], story: Dict[str, Any],
                 next_fr_num: int = 200, next_tc_prefix: str = "GENERATED") -> Tuple[str, str]:
    """
    Format a prompt template with the story's fields.

    Returns (system_prompt, user_prompt).
    """
    frs_text = story.get("frs", "Not provided.")
    accept = story.get("acceptance_criteria", "")
    deps = story.get("dependencies", "None")

    user_text = template["user_prompt_template"].format(
        next_fr_num=next_fr_num,
        next_tc_prefix=next_tc_prefix,
        title=story.get("title", ""),
        domain=story.get("domain", ""),
        priority=story.get("priority", ""),
        user_story=story.get("user_story", ""),
        acceptance_criteria=accept,
        dependencies=deps,
        frs=frs_text,
    )
    # Prepend story ID so backends can identify the source story
    user_text = f"[Story: {story.get('story_id', 'UNKNOWN')}]\n\n{user_text}"
    return template["system_prompt"], user_text


def compute_prompt_hash(system_prompt: str, user_prompt: str) -> str:
    """Short SHA-256 digest of the prompt pair (for tracking)."""
    combined = f"{system_prompt}\n---\n{user_prompt}"
    return hashlib.sha256(combined.encode("utf-8")).hexdigest()[:12]


def generate_for_story(
    story: Dict[str, Any],
    frs_template_name: str,
    tc_template_name: str,
    backend: LlmBackend,
    base_fr_num: int,
    base_tc_prefix: str,
) -> Dict[str, Any]:
    """
    Generate FRS and test cases for a single user story using the
    specified templates and LLM backend. Returns a result record.
    """
    frs_tmpl = PROMPT_TEMPLATES[frs_template_name]
    tc_tmpl = PROMPT_TEMPLATES[tc_template_name]

    sys_frs, usr_frs = build_prompt(frs_tmpl, story, next_fr_num=base_fr_num)
    sys_tc, usr_tc = build_prompt(tc_tmpl, story, next_tc_prefix=base_tc_prefix)

    frs_hash = compute_prompt_hash(sys_frs, usr_frs)
    tc_hash = compute_prompt_hash(sys_tc, usr_tc)

    timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    story_id = story.get("story_id", "UNKNOWN")

    try:
        generated_frs = backend.complete(sys_frs, usr_frs,
                                         metadata={"story_id": story_id, "category": frs_template_name})
    except Exception as exc:
        generated_frs = f"ERROR: {exc}"

    try:
        generated_tc_raw = backend.complete(sys_tc, usr_tc,
                                            metadata={"story_id": story_id, "category": tc_template_name})
    except Exception as exc:
        generated_tc_raw = f"ERROR: {exc}"

    # Try to parse TC JSON
    parsed_tcs: List[Dict[str, str]] = []
    try:
        parsed_raw = json.loads(generated_tc_raw)
        if isinstance(parsed_raw, list):
            parsed_tcs = parsed_raw
        else:
            parsed_tcs = [{
                "tc_id": f"TC-{base_tc_prefix}-01",
                "description": "LLM returned non-list — see raw output",
                "test_type": "N/A",
                "steps": f"LLM returned type {type(parsed_raw).__name__}. Refer to raw output.",
                "expected": generated_tc_raw[:500],
            }]
    except json.JSONDecodeError:
        parsed_tcs = [{
            "tc_id": f"TC-{base_tc_prefix}-01",
            "description": "Parsing failed — see raw output",
            "test_type": "N/A",
            "steps": "Refer to raw generated output.",
            "expected": generated_tc_raw[:500],
        }]

    n_fr_lines = len([l for l in generated_frs.split("\n") if l.strip().startswith("FR-")])

    return {
        "story_id": story["story_id"],
        "title": story["title"],
        "domain": story["domain"],
        "priority": story["priority"],
        "user_story": story["user_story"],
        "acceptance_criteria": story["acceptance_criteria"],
        # Prompt info
        "frs_prompt_name": frs_template_name,
        "frs_prompt_hash": frs_hash,
        "tc_prompt_name": tc_template_name,
        "tc_prompt_hash": tc_hash,
        # LLM metadata
        "llm_provider": backend.provider_name,
        "llm_model": backend.model_name,
        "generation_timestamp": timestamp,
        # Generated content
        "generated_frs": generated_frs,
        "generated_frs_line_count": n_fr_lines,
        "generated_tc_raw": generated_tc_raw,
        "generated_tc_parsed": parsed_tcs,
        "generated_tc_count": len(parsed_tcs),
    }


# ═══════════════════════════════════════════════════════════════════════════
# 5. Excel Writer
# ═══════════════════════════════════════════════════════════════════════════

STYLES = {
    "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
    "header_fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "cell_font": Font(name="Calibri", size=10),
    "cell_align": Alignment(vertical="top", wrap_text=True),
    "cell_align_center": Alignment(horizontal="center", vertical="top", wrap_text=True),
    "thin_border": Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    ),
    "alt_fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "green_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

def _apply_header_style(cell):
    cell.font = STYLES["header_font"]
    cell.fill = STYLES["header_fill"]
    cell.alignment = STYLES["header_align"]
    cell.border = STYLES["thin_border"]

def _apply_cell_style(cell, center: bool = False, alt: bool = False):
    cell.font = STYLES["cell_font"]
    cell.alignment = STYLES["cell_align_center"] if center else STYLES["cell_align"]
    cell.border = STYLES["thin_border"]
    if alt:
        cell.fill = STYLES["alt_fill"]


def append_generated_results_to_excel(
    excel_path: str,
    results: List[Dict[str, Any]],
    prompts_used: List[str],
):
    """
    Load the existing workbook, add/edit sheets for generated results,
    prompt templates, and run metadata. Saves in place.
    """
    wb: openpyxl.Workbook
    if Path(excel_path).exists():
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()

    # ── Sheet: Generated FRS & Test Cases ────────────────────────────────
    if "Generated Results" in wb.sheetnames:
        del wb["Generated Results"]
    ws_r = wb.create_sheet("Generated Results", 0)

    headers = [
        "Story ID", "Title", "Domain", "Priority",
        "FRS Prompt", "FRS Prompt Hash", "TC Prompt", "TC Prompt Hash",
        "LLM Provider", "LLM Model", "Generation Timestamp",
        "Generated FRS Count", "Generated TC Count",
        "Generated FRS", "Generated TC Raw", "Generated TC Parsed",
        "User Story", "Acceptance Criteria",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws_r.cell(row=1, column=c, value=h)
        _apply_header_style(cell)

    for r, rec in enumerate(results, 2):
        values = [
            rec["story_id"], rec["title"], rec["domain"], rec["priority"],
            rec["frs_prompt_name"], rec["frs_prompt_hash"],
            rec["tc_prompt_name"], rec["tc_prompt_hash"],
            rec["llm_provider"], rec["llm_model"], rec["generation_timestamp"],
            rec["generated_frs_line_count"], rec["generated_tc_count"],
            rec["generated_frs"], rec["generated_tc_raw"],
            json.dumps(rec["generated_tc_parsed"], indent=2),
            rec["user_story"], rec["acceptance_criteria"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws_r.cell(row=r, column=c, value=v)
            _apply_cell_style(cell, center=c <= 4, alt=(r % 2 == 0))

    col_widths = [12, 35, 18, 10, 20, 16, 20, 16, 14, 22, 22, 16, 16, 70, 70, 60, 60, 60]
    for i, w in enumerate(col_widths, 1):
        ws_r.column_dimensions[get_column_letter(i)].width = w
    ws_r.auto_filter.ref = ws_r.dimensions
    ws_r.freeze_panes = "A2"

    # ── Sheet: Prompt Templates ──────────────────────────────────────────
    if "Prompt Templates" in wb.sheetnames:
        del wb["Prompt Templates"]
    ws_p = wb.create_sheet("Prompt Templates")

    tmpl_headers = ["Template Name", "Description", "Category", "System Prompt", "User Prompt Template"]
    for c, h in enumerate(tmpl_headers, 1):
        cell = ws_p.cell(row=1, column=c, value=h)
        _apply_header_style(cell)

    row = 2
    for name in prompts_used:  # write only the ones used, but all exist in PROMPT_TEMPLATES
        tmpl = PROMPT_TEMPLATES.get(name)
        if tmpl is None:
            continue
        vals = [name, tmpl["description"], tmpl["category"],
                tmpl["system_prompt"], tmpl["user_prompt_template"]]
        for c, v in enumerate(vals, 1):
            cell = ws_p.cell(row=row, column=c, value=v)
            _apply_cell_style(cell, alt=(row % 2 == 0))
        row += 1
    # Also write all other templates not used yet
    for name, tmpl in PROMPT_TEMPLATES.items():
        if name in prompts_used:
            continue
        vals = [name, tmpl["description"], tmpl["category"],
                tmpl["system_prompt"], tmpl["user_prompt_template"]]
        for c, v in enumerate(vals, 1):
            cell = ws_p.cell(row=row, column=c, value=v)
            _apply_cell_style(cell, alt=(row % 2 == 0))
        row += 1

    widths_p = [20, 60, 14, 80, 100]
    for i, w in enumerate(widths_p, 1):
        ws_p.column_dimensions[get_column_letter(i)].width = w
    ws_p.auto_filter.ref = ws_p.dimensions
    ws_p.freeze_panes = "A2"

    # ── Sheet: Generated FRS (Normalized) ─────────────────────────────────
    if "Generated FRS" in wb.sheetnames:
        del wb["Generated FRS"]
    ws_gf = wb.create_sheet("Generated FRS")

    gf_headers = ["Story ID", "Title", "FR Reference", "Generated Functional Requirement"]
    for c, h in enumerate(gf_headers, 1):
        cell = ws_gf.cell(row=1, column=c, value=h)
        _apply_header_style(cell)

    gf_row = 2
    for rec in results:
        fr_lines = [l.strip() for l in rec["generated_frs"].split("\n") if l.strip()]
        for line in fr_lines:
            fr_ref = ""
            fr_text = line
            if ":" in line and (line.strip().startswith("FR-") or line.strip().startswith("ERROR")):
                parts = line.split(":", 1)
                fr_ref = parts[0].strip()
                fr_text = parts[1].strip() if len(parts) > 1 else line
            vals = [rec["story_id"], rec["title"], fr_ref, fr_text]
            for c, v in enumerate(vals, 1):
                cell = ws_gf.cell(row=gf_row, column=c, value=v)
                _apply_cell_style(cell, center=c <= 2, alt=(gf_row % 2 == 0))
            gf_row += 1

    widths_gf = [12, 35, 16, 100]
    for i, w in enumerate(widths_gf, 1):
        ws_gf.column_dimensions[get_column_letter(i)].width = w
    ws_gf.auto_filter.ref = ws_gf.dimensions
    ws_gf.freeze_panes = "A2"

    # ── Sheet: Generated TCs (Normalized) ─────────────────────────────────
    if "Generated TCs" in wb.sheetnames:
        del wb["Generated TCs"]
    ws_gt = wb.create_sheet("Generated TCs")

    gt_headers = ["Story ID", "Title", "Test Case ID", "Test Description",
                  "Test Type", "Test Steps", "Expected Result"]
    for c, h in enumerate(gt_headers, 1):
        cell = ws_gt.cell(row=1, column=c, value=h)
        _apply_header_style(cell)

    gt_row = 2
    for rec in results:
        for tc in rec.get("generated_tc_parsed", []):
            vals = [
                rec["story_id"], rec["title"],
                tc.get("tc_id", ""), tc.get("description", ""),
                tc.get("test_type", ""), tc.get("steps", ""), tc.get("expected", ""),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws_gt.cell(row=gt_row, column=c, value=v)
                _apply_cell_style(cell, center=c <= 3, alt=(gt_row % 2 == 0))
            gt_row += 1

    widths_gt = [12, 35, 16, 50, 14, 60, 60]
    for i, w in enumerate(widths_gt, 1):
        ws_gt.column_dimensions[get_column_letter(i)].width = w
    ws_gt.auto_filter.ref = ws_gt.dimensions
    ws_gt.freeze_panes = "A2"

    # ── Sheet: Run Metadata ──────────────────────────────────────────────
    if "Run Metadata" in wb.sheetnames:
        del wb["Run Metadata"]
    ws_m = wb.create_sheet("Run Metadata")

    meta_items = [
        ("Script Version", "1.0"),
        ("Generated At", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Total Stories Processed", len(results)),
        ("Prompt Categories", f"frs ({len([p for p in PROMPT_TEMPLATES if PROMPT_TEMPLATES[p]['category']=='frs'])}), "
                             f"test_case ({len([p for p in PROMPT_TEMPLATES if PROMPT_TEMPLATES[p]['category']=='test_case'])}))"),
        ("Backend Used", results[0]["llm_provider"] if results else "N/A"),
        ("Model Used", results[0]["llm_model"] if results else "N/A"),
        ("Prompt Strategy (FRS)", prompts_used[0] if prompts_used else "N/A"),
        ("Prompt Strategy (TC)", prompts_used[1] if len(prompts_used) > 1 else "N/A"),
        ("", ""),
        ("## Extensibility Notes", ""),
        ("New stories source", "Use --stories JSON flag to import additional user stories from a JSON file"),
        ("New LLM backends", "Subclass LlmBackend; register in BACKEND_REGISTRY"),
        ("New prompts", "Add entries to PROMPT_TEMPLATES dict"),
        ("Output format", "All results in this workbook with prompt hashes for traceability"),
    ]
    for r, (label, val) in enumerate(meta_items, 1):
        c1 = ws_m.cell(row=r, column=1, value=label)
        c2 = ws_m.cell(row=r, column=2, value=val)
        c1.font = Font(name="Calibri", size=10, bold=label.startswith("##") or not label.startswith("  "))
        c2.font = STYLES["cell_font"]
        c1.alignment = STYLES["cell_align"]
        c2.alignment = STYLES["cell_align"]
        c1.border = STYLES["thin_border"]
        c2.border = STYLES["thin_border"]

    ws_m.column_dimensions["A"].width = 35
    ws_m.column_dimensions["B"].width = 80

    wb.save(excel_path)
    print(f"Workbook saved to {excel_path}")


# ═══════════════════════════════════════════════════════════════════════════
# 6. Main Entry Point
# ═══════════════════════════════════════════════════════════════════════════

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate FRS and test cases via LLM prompts"
    )
    parser.add_argument("--output", default=None,
                        help="Path to output Excel file (default: Dataset_...xlsx)")
    parser.add_argument("--backend", default="internal",
                        choices=list(BACKEND_REGISTRY),
                        help="LLM backend to use")
    parser.add_argument("--model", default=None,
                        help="Model name for the backend (e.g. gpt-4o, claude-sonnet-4-20250514)")
    parser.add_argument("--frs-prompt", default="frs_few_shot",
                        help="Prompt template key for FRS generation")
    parser.add_argument("--tc-prompt", default="tc_few_shot",
                        help="Prompt template key for test case generation")
    parser.add_argument("--stories", default=None,
                        help="Optional JSON file with additional user stories")
    parser.add_argument("--limit", type=int, default=5,
                        help="Max number of stories to process (default 5, for demo)")
    args = parser.parse_args()

    # Resolve output path
    base_dir = Path(__file__).resolve().parent
    output_path = str(base_dir / (args.output or "Dataset_UserStories_FRS_TestCases.xlsx"))

    # Load stories
    stories = load_user_stories()
    if args.stories:
        with open(args.stories) as fh:
            extra = json.load(fh)
            stories.extend(extra)
    stories = stories[: args.limit]

    # Initialize backend
    backend_cls = BACKEND_REGISTRY[args.backend]
    backend_kwargs: Dict[str, Any] = {}
    if args.model:
        backend_kwargs["model"] = args.model
    backend = backend_cls(**backend_kwargs)
    print(f"Using backend: {backend.provider_name} / {backend.model_name}")

    # Validate prompt template names
    if args.frs_prompt not in PROMPT_TEMPLATES:
        print(f"Unknown FRS prompt: {args.frs_prompt}. Available: {[k for k in PROMPT_TEMPLATES if PROMPT_TEMPLATES[k]['category']=='frs']}")
        sys.exit(1)
    if args.tc_prompt not in PROMPT_TEMPLATES:
        print(f"Unknown TC prompt: {args.tc_prompt}. Available: {[k for k in PROMPT_TEMPLATES if PROMPT_TEMPLATES[k]['category']=='test_case']}")
        sys.exit(1)

    # Generate
    results: List[Dict[str, Any]] = []
    base_fr = 200  # Starting FR number for generated content
    for i, story in enumerate(stories):
        print(f"  Processing {story['story_id']}: {story['title'][:60]}...")
        story_num = story["story_id"].replace("US-", "")
        tc_prefix = f"{story_num}-G"
        rec = generate_for_story(
            story=story,
            frs_template_name=args.frs_prompt,
            tc_template_name=args.tc_prompt,
            backend=backend,
            base_fr_num=base_fr + i * 20,  # 20 FR slots per story
            base_tc_prefix=tc_prefix,
        )
        results.append(rec)

    # Write the Excel
    prompts_used = [args.frs_prompt, args.tc_prompt]
    append_generated_results_to_excel(output_path, results, prompts_used)

    # Summary
    total_frs = sum(r["generated_frs_line_count"] for r in results)
    total_tc = sum(r["generated_tc_count"] for r in results)
    print(f"\nDone. {len(results)} stories processed.")
    print(f"  Total generated FRs: {total_frs}")
    print(f"  Total generated TCs: {total_tc}")
    print(f"  Output: {output_path}")
    print(f"\nTip: run with --backend openai --model gpt-4o for real LLM generation")

    # ── Also verify the Excel is readable ─────────────────────────────────
    try:
        wb_check = openpyxl.load_workbook(output_path)
        print(f"  Sheets: {wb_check.sheetnames}")
        wb_check.close()
    except Exception as exc:
        print(f"  Warning: could not verify workbook: {exc}")


if __name__ == "__main__":
    main()
