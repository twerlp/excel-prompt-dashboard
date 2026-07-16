"""
Generate an Excel dataset of user stories with their corresponding
Functional Requirement Specifications (FRS) and test cases.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ── dataset ──────────────────────────────────────────────────────────────
stories = [
    # 1
    {
        "story_id": "US-001",
        "title": "User Registration with Email Verification",
        "domain": "E-Commerce",
        "priority": "High",
        "story_points": 8,
        "user_story": "As a new customer, I want to register an account using my email and verify it, so that I can securely access the platform and track my orders.",
        "acceptance_criteria": (
            "1. User enters name, email, password (min 8 chars, 1 uppercase, 1 digit) on registration form.\n"
            "2. System validates email format and checks for duplicate accounts.\n"
            "3. On success, a 6-digit verification code is sent to the email within 60 seconds.\n"
            "4. Code expires after 10 minutes; user may request resend up to 3 times.\n"
            "5. Account activates only after correct code entry.\n"
            "6. Password is stored using bcrypt hashing."
        ),
        "frs": (
            "FR-001: The system shall provide a self-service registration form capturing full name, "
            "email address, and password.\n"
            "FR-002: The system shall enforce password complexity rules — minimum 8 characters, "
            "at least 1 uppercase letter and 1 digit.\n"
            "FR-003: The system shall validate the email address against RFC 5322 format and "
            "reject duplicates against the user store.\n"
            "FR-004: The system shall generate a cryptographically random 6-digit numeric "
            "verification token with a 10-minute TTL.\n"
            "FR-005: The system shall send the token via an SMTP email service within 60 seconds "
            "of account creation.\n"
            "FR-006: The system shall allow up to 3 token resend requests per registration session.\n"
            "FR-007: The system shall hash and salt passwords using bcrypt (cost factor ≥ 10) "
            "before persisting to the database.\n"
            "FR-008: The system shall activate the account and create a default user profile "
            "upon successful token verification."
        ),
        "test_cases": [
            {
                "tc_id": "TC-001-01",
                "description": "Verify successful registration with valid input and email verification.",
                "test_type": "Functional",
                "steps": (
                    "1. Navigate to registration page.\n"
                    "2. Enter valid name, unique email, compliant password.\n"
                    "3. Submit form.\n"
                    "4. Retrieve verification code from test email inbox.\n"
                    "5. Enter code on verification screen."
                ),
                "expected": "Account is created in 'pending' state; after code entry state → 'active'. User can log in.",
            },
            {
                "tc_id": "TC-001-02",
                "description": "Verify duplicate email is rejected.",
                "test_type": "Functional",
                "steps": (
                    "1. Register with email already in use.\n"
                    "2. Submit form."
                ),
                "expected": "System displays error 'An account with this email already exists.' No second record created.",
            },
            {
                "tc_id": "TC-001-03",
                "description": "Verify weak password is rejected.",
                "test_type": "Functional",
                "steps": (
                    "1. Enter password 'abc' (too short, no uppercase, no digit).\n"
                    "2. Submit form."
                ),
                "expected": "System displays validation error listing each unmet rule.",
            },
            {
                "tc_id": "TC-001-04",
                "description": "Verify expired token is rejected.",
                "test_type": "Functional",
                "steps": (
                    "1. Wait 10+ minutes after registration before entering token.\n"
                    "2. Enter correct but expired token."
                ),
                "expected": "System displays 'Verification code has expired. Request a new one.'",
            },
            {
                "tc_id": "TC-001-05",
                "description": "Verify resend limit enforcement.",
                "test_type": "Functional",
                "steps": (
                    "1. Click 'Resend Code' 4 times.\n"
                    "2. Observe behavior on 4th click."
                ),
                "expected": "First 3 resends succeed; 4th is blocked with message 'Maximum resend attempts reached.'",
            },
            {
                "tc_id": "TC-001-06",
                "description": "Verify password is stored as hash, not plaintext.",
                "test_type": "Security",
                "steps": (
                    "1. Register a user.\n"
                    "2. Query the database directly for the password field."
                ),
                "expected": "Stored value is a bcrypt hash string (prefix $2b$), not plaintext.",
            },
        ],
        "dependencies": "SMTP email service; user database schema",
        "sprint": "Sprint 2",
        "status": "Done",
    },
    # 2
    {
        "story_id": "US-002",
        "title": "Shopping Cart — Add, Update, Remove Items",
        "domain": "E-Commerce",
        "priority": "High",
        "story_points": 13,
        "user_story": "As a shopper, I want to add products to a cart, change quantities, and remove items, so that I can manage my order before checkout.",
        "acceptance_criteria": (
            "1. Guest users get a session-based cart; logged-in users have a persistent cart.\n"
            "2. Adding an item increments quantity if it already exists in the cart.\n"
            "3. Quantity can be updated via +/- controls or direct input (1 to stock max).\n"
            "4. Removing an item shows a confirmation toast.\n"
            "5. Cart subtotal updates in real-time on any change.\n"
            "6. Cart icon in header shows item count badge."
        ),
        "frs": (
            "FR-009: The system shall maintain a cart data structure keyed by anonymous session ID "
            "for guest users and by user ID for authenticated users.\n"
            "FR-010: The system shall merge a guest cart into the user's persistent cart upon login.\n"
            "FR-011: The system shall provide REST endpoints: POST /cart/items, "
            "PATCH /cart/items/{sku}, DELETE /cart/items/{sku}.\n"
            "FR-012: Add-item endpoint shall accept SKU and quantity; if SKU already present, "
            "quantity shall be incremented.\n"
            "FR-013: Quantity constraints — minimum 1 (triggers removal prompt if set to 0), "
            "maximum the current inventory count for the SKU.\n"
            "FR-014: The system shall recalculate the cart subtotal (unit price × quantity) "
            "server-side after each mutation and return it to the client.\n"
            "FR-015: The header cart badge shall display the total unique item count via a client-side "
            "subscription or polling mechanism."
        ),
        "test_cases": [
            {
                "tc_id": "TC-002-01",
                "description": "Verify adding a new item to an empty cart.",
                "test_type": "Functional",
                "steps": "1. Browse product listing. 2. Click 'Add to Cart' for SKU A. 3. Open cart drawer.",
                "expected": "Cart contains 1 × SKU A. Subtotal equals unit price. Badge shows '1'.",
            },
            {
                "tc_id": "TC-002-02",
                "description": "Verify adding an existing item increments quantity.",
                "test_type": "Functional",
                "steps": "1. Cart has 1 × SKU A. 2. Add SKU A again.",
                "expected": "Cart shows 2 × SKU A. Subtotal = 2 × unit price.",
            },
            {
                "tc_id": "TC-002-03",
                "description": "Verify quantity update via direct input.",
                "test_type": "Functional",
                "steps": "1. Change quantity of SKU A from 1 to 5 via input field. 2. Blur field.",
                "expected": "Cart updates to 5 × SKU A; subtotal recalculated.",
            },
            {
                "tc_id": "TC-002-04",
                "description": "Verify quantity cannot exceed inventory.",
                "test_type": "Boundary",
                "steps": "1. Inventory for SKU B = 10. 2. Enter quantity 11. 3. Blur field.",
                "expected": "Quantity snaps back to 10. Warning toast: 'Only 10 units available.'",
            },
            {
                "tc_id": "TC-002-05",
                "description": "Verify item removal with confirmation.",
                "test_type": "Functional",
                "steps": "1. Click remove icon on a cart item. 2. Confirm in toast. 3. Observe cart.",
                "expected": "Item removed; subtotal recalculated; badge count decremented.",
            },
            {
                "tc_id": "TC-002-06",
                "description": "Verify guest cart merges on login.",
                "test_type": "Functional",
                "steps": "1. As guest, add SKU C to cart. 2. Log in to account that has SKU C already. 3. Check cart.",
                "expected": "Cart shows 2 × SKU C (guest qty + logged-in qty).",
            },
        ],
        "dependencies": "Product catalog service; inventory service; authentication service",
        "sprint": "Sprint 3",
        "status": "Done",
    },
    # 3
    {
        "story_id": "US-003",
        "title": "Secure Payment Processing via Stripe",
        "domain": "E-Commerce",
        "priority": "Critical",
        "story_points": 21,
        "user_story": "As a customer, I want to pay with my credit card securely through the platform, so that I can complete my purchase with confidence.",
        "acceptance_criteria": (
            "1. User selects stored card or enters new card details in a Stripe Elements iframe.\n"
            "2. Card data never touches backend servers (PCI-DSS compliance).\n"
            "3. Payment succeeds — order status → 'Paid', inventory reduced.\n"
            "4. Payment fails (insufficient funds / expired card) — user sees specific error.\n"
            "5. Payment is idempotent; duplicate submission handles gracefully.\n"
            "6. Confirmation email with receipt is sent automatically."
        ),
        "frs": (
            "FR-016: The system shall integrate Stripe Elements for client-side tokenized card collection, "
            "ensuring no raw PAN (Primary Account Number) is transmitted to application servers.\n"
            "FR-017: The system shall implement a /checkout API endpoint that accepts a Stripe PaymentMethod ID "
            "and order ID, creates a PaymentIntent via Stripe SDK, and returns the client secret for confirmation.\n"
            "FR-018: The server shall use Stripe webhooks (payment_intent.succeeded, payment_intent.payment_failed) "
            "to update order status asynchronously.\n"
            "FR-019: The system shall enforce idempotency using an idempotency key derived from order ID, "
            "preventing duplicate charges.\n"
            "FR-020: On successful payment, the system shall decrement inventory atomically within a database "
            "transaction and mark the order as 'Paid'.\n"
            "FR-021: The system shall send a transactional email containing order summary and receipt link "
            "via the notification service.\n"
            "FR-022: The system shall log all payment events in a PCI-compliant audit table, excluding "
            "sensitive card data."
        ),
        "test_cases": [
            {
                "tc_id": "TC-003-01",
                "description": "Verify successful payment with a valid test card.",
                "test_type": "Functional",
                "steps": "1. Add items to cart → checkout. 2. Enter Stripe test card 4242... 3. Confirm payment.",
                "expected": "Order status → 'Paid'. Inventory decremented. Confirmation email received. Redirect to order success page.",
            },
            {
                "tc_id": "TC-003-02",
                "description": "Verify payment failure with declined card.",
                "test_type": "Functional",
                "steps": "1. Use Stripe test card 4000000000000002 (declined). 2. Submit payment.",
                "expected": "Error message 'Your card was declined.' Order remains 'Pending'. No inventory change.",
            },
            {
                "tc_id": "TC-003-03",
                "description": "Verify idempotency — double-click protection.",
                "test_type": "Functional",
                "steps": "1. Submit payment. 2. Intercept and resend the same request with identical idempotency key before first completes.",
                "expected": "Only one PaymentIntent created. Both responses return same outcome. Only one charge appears in Stripe dashboard.",
            },
            {
                "tc_id": "TC-003-04",
                "description": "Verify webhook handling when payment_intent.succeeded fires.",
                "test_type": "Integration",
                "steps": "1. Trigger a successful test payment. 2. Wait for Stripe webhook delivery.",
                "expected": "Webhook received → order status updated → email dispatched → webhook returns HTTP 200 to Stripe.",
            },
            {
                "tc_id": "TC-003-05",
                "description": "Verify no card data in application logs or database.",
                "test_type": "Security",
                "steps": "1. Complete a payment. 2. Inspect application logs and orders table.",
                "expected": "No PAN, CVC, or expiry date present in logs or database columns.",
            },
        ],
        "dependencies": "Stripe account and API keys; webhook endpoint; notification service",
        "sprint": "Sprint 4",
        "status": "In Progress",
    },
    # 4
    {
        "story_id": "US-004",
        "title": "Search Products with Filters and Pagination",
        "domain": "E-Commerce",
        "priority": "Medium",
        "story_points": 8,
        "user_story": "As a shopper, I want to search products by keyword and filter by category, price range, and rating, so that I can quickly find relevant items.",
        "acceptance_criteria": (
            "1. Search bar accepts text; results appear in ≤ 500 ms for catalog ≤ 100k products.\n"
            "2. Category filter is a multi-select dropdown populated from the catalog.\n"
            "3. Price range is entered as min/max with client-side validation (min ≤ max, non-negative).\n"
            "4. Rating filter uses 1–5 star toggle buttons.\n"
            "5. Results are paginated with 20 items per page; page navigation shows up to 5 page links.\n"
            "6. 'No results' state shows suggestions and a link to clear all filters."
        ),
        "frs": (
            "FR-023: The system shall provide a GET /products endpoint accepting query params: "
            "q (text), category (comma-separated), min_price, max_price, min_rating, page, and limit.\n"
            "FR-024: The system shall perform full-text search against product name and description "
            "columns using PostgreSQL tsvector/tsquery or equivalent indexed search.\n"
            "FR-025: The system shall return results with metadata: total_count, page, total_pages, "
            "and an array of product objects (id, name, price, rating, thumbnail_url).\n"
            "FR-026: The system shall cache popular search results in Redis with a 5-minute TTL.\n"
            "FR-027: The client shall debounce search input by 300 ms before firing the API request.\n"
            "FR-028: The category filter options shall be fetched from GET /categories and rendered "
            "as a dynamic multi-select component."
        ),
        "test_cases": [
            {
                "tc_id": "TC-004-01",
                "description": "Verify keyword search returns matching products.",
                "test_type": "Functional",
                "steps": "1. Enter 'wireless headphones' in search bar. 2. Wait for results.",
                "expected": "Results contain products whose name or description includes 'wireless' and 'headphones'. Relevance-sorted.",
            },
            {
                "tc_id": "TC-004-02",
                "description": "Verify combined filters narrow results.",
                "test_type": "Functional",
                "steps": "1. Search 'laptop'. 2. Select category 'Electronics'. 3. Set price $500–$1500. 4. Set rating 4★+.",
                "expected": "Results are Electronics laptops priced $500–$1500 with average rating ≥ 4.0.",
            },
            {
                "tc_id": "TC-004-03",
                "description": "Verify pagination navigation.",
                "test_type": "Functional",
                "steps": "1. Perform search that returns 45 results. 2. Observe page 1 (20 items). 3. Click page 2 and page 3.",
                "expected": "Page 1: items 1–20. Page 2: 21–40. Page 3: 41–45. URL reflects page param.",
            },
            {
                "tc_id": "TC-004-04",
                "description": "Verify empty results state.",
                "test_type": "Functional",
                "steps": "1. Search 'unicorn-teleporter'. 2. Observe results area.",
                "expected": "Display 'No products found' with suggestion text and 'Clear Filters' link. Link resets all params.",
            },
            {
                "tc_id": "TC-004-05",
                "description": "Verify response time SLA.",
                "test_type": "Performance",
                "steps": "1. Seed catalog with 100k products. 2. Run 100 concurrent search requests via JMeter/k6.",
                "expected": "p95 latency ≤ 500 ms; p99 ≤ 800 ms.",
            },
        ],
        "dependencies": "Product catalog service; PostgreSQL full-text search index; Redis cache",
        "sprint": "Sprint 2",
        "status": "Done",
    },
    # 5
    {
        "story_id": "US-005",
        "title": "Role-Based Access Control for Admin Dashboard",
        "domain": "Enterprise SaaS",
        "priority": "High",
        "story_points": 13,
        "user_story": "As a system administrator, I want to assign roles (Admin, Editor, Viewer) to users and control their permissions, so that sensitive operations are restricted to authorized personnel.",
        "acceptance_criteria": (
            "1. Admin can create, update, and delete roles from the Roles management page.\n"
            "2. Each role defines granular permissions: create, read, update, delete per resource.\n"
            "3. Admin assigns roles to users; users inherit the permissions of their highest role.\n"
            "4. UI elements (buttons, links, pages) are conditionally rendered based on permissions.\n"
            "5. Backend middleware enforces permission checks on every protected API route.\n"
            "6. Unauthorized access returns HTTP 403; audit event is logged."
        ),
        "frs": (
            "FR-029: The system shall implement RBAC with three entities: User, Role, Permission.\n"
            "FR-030: Permissions shall be defined as 'resource:action' tuples (e.g., 'user:create', "
            "'report:read') stored in a dedicated permission table.\n"
            "FR-031: A Role shall be a named collection of permissions with a many-to-many relationship "
            "between roles and permissions.\n"
            "FR-032: A User shall have a many-to-many relationship with roles; effective permissions "
            "shall be the union of all assigned roles' permissions.\n"
            "FR-033: The API gateway/middleware shall intercept every authenticated request, resolve "
            "the user's effective permissions, and compare against the required permission for the endpoint.\n"
            "FR-034: The frontend shall fetch the user's permission set on login and use a directive "
            "(e.g., v-can) to conditionally render UI elements.\n"
            "FR-035: Every denied access attempt shall be logged to an audit log with user ID, "
            "resource, action, timestamp, and IP address."
        ),
        "test_cases": [
            {
                "tc_id": "TC-005-01",
                "description": "Verify Viewer cannot access admin pages.",
                "test_type": "Functional",
                "steps": "1. Log in as Viewer. 2. Attempt to navigate to /admin/users via URL.",
                "expected": "HTTP 403 returned. Redirected to 'Access Denied' page. Audit log entry created.",
            },
            {
                "tc_id": "TC-005-02",
                "description": "Verify Admin can create a new role.",
                "test_type": "Functional",
                "steps": "1. Log in as Admin. 2. Navigate to Roles page. 3. Click 'New Role'. 4. Name it 'Auditor', select 'report:read', 'user:read'. 5. Save.",
                "expected": "Role 'Auditor' appears in role list. Permissions persisted.",
            },
            {
                "tc_id": "TC-005-03",
                "description": "Verify UI elements hidden for unauthorized role.",
                "test_type": "Functional",
                "steps": "1. Log in as Viewer. 2. Navigate to Reports page.",
                "expected": "'Delete Report' button not rendered; 'Export CSV' button visible (report:read).",
            },
            {
                "tc_id": "TC-005-04",
                "description": "Verify permission change propagates immediately.",
                "test_type": "Functional",
                "steps": "1. Admin upgrades Viewer → Editor. 2. Viewer refreshes or re-logs in.",
                "expected": "Editor now sees 'Create Report' button; can POST to /api/reports.",
            },
        ],
        "dependencies": "Authentication service; audit logging service",
        "sprint": "Sprint 5",
        "status": "In Progress",
    },
    # 6
    {
        "story_id": "US-006",
        "title": "Funds Transfer Between Accounts",
        "domain": "Banking",
        "priority": "Critical",
        "story_points": 21,
        "user_story": "As a bank customer, I want to transfer money between my own accounts or to another beneficiary, so that I can manage my finances without visiting a branch.",
        "acceptance_criteria": (
            "1. User selects source account, destination account/beneficiary, amount, and optional note.\n"
            "2. Transfer to self is instant; transfer to external beneficiary uses scheduled batch (NEFT/ACH).\n"
            "3. Balance validation — source must have sufficient funds plus any applicable fees.\n"
            "4. Daily transfer limit of $50,000 enforced; configurable per customer tier.\n"
            "5. Two-factor authentication required for transfers > $5,000.\n"
            "6. Transaction receipt downloadable as PDF."
        ),
        "frs": (
            "FR-036: The system shall provide a /transfer endpoint accepting source_account_id, "
            "destination_account_id (or beneficiary_id), amount, currency (default USD), and note.\n"
            "FR-037: The system shall execute intra-bank transfers synchronously within a database "
            "transaction — debit source, credit destination, insert transaction record.\n"
            "FR-038: The system shall enqueue external (inter-bank) transfers to the NEFT/ACH batch "
            "processing pipeline for settlement during the next clearing window.\n"
            "FR-039: The system shall validate available balance (including hold amounts) before "
            "allowing debit.\n"
            "FR-040: The system shall enforce a rolling 24-hour cumulative transfer limit per account, "
            "configurable by customer tier in a feature-flag service.\n"
            "FR-041: For transfers exceeding the 2FA threshold ($5,000 default), the system shall "
            "trigger a secondary authentication challenge (OTP or push notification).\n"
            "FR-042: The system shall generate a PDF receipt containing transaction reference, "
            "date, amounts, and account masked details, served via GET /transactions/{id}/receipt."
        ),
        "test_cases": [
            {
                "tc_id": "TC-006-01",
                "description": "Verify successful intra-bank transfer.",
                "test_type": "Functional",
                "steps": "1. Select savings (balance $10,000) → checking (balance $500). 2. Transfer $1,000. 3. Submit.",
                "expected": "Savings → $9,000. Checking → $1,500. Transaction record created with status 'Completed'. Receipt available.",
            },
            {
                "tc_id": "TC-006-02",
                "description": "Verify transfer blocked for insufficient funds.",
                "test_type": "Functional",
                "steps": "1. Savings balance $500. 2. Attempt transfer $1,000. 3. Submit.",
                "expected": "Error: 'Insufficient funds. Available balance: $500.00.' No debit recorded.",
            },
            {
                "tc_id": "TC-006-03",
                "description": "Verify daily transfer limit enforcement.",
                "test_type": "Functional",
                "steps": "1. Transfer $49,000 (within $50k limit). 2. Attempt second transfer of $2,000 on same day.",
                "expected": "First transfer succeeds. Second blocked with 'Daily transfer limit exceeded. Remaining: $1,000.'",
            },
            {
                "tc_id": "TC-006-04",
                "description": "Verify 2FA trigger for high-value transfer.",
                "test_type": "Functional",
                "steps": "1. Initiate transfer of $6,000 (above $5,000 threshold). 2. Submit.",
                "expected": "System prompts for OTP or push notification approval. Transfer only proceeds after 2FA success.",
            },
            {
                "tc_id": "TC-006-05",
                "description": "Verify transaction atomicity — system crash mid-transfer.",
                "test_type": "Reliability",
                "steps": "1. Initiate intra-bank transfer. 2. Kill database connection after debit but before credit.",
                "expected": "Transaction rolled back. Both accounts retain original balances. No orphan transaction record.",
            },
        ],
        "dependencies": "Core banking system; NEFT/ACH gateway; OTP service; PDF generation service",
        "sprint": "Sprint 6",
        "status": "Backlog",
    },
    # 7
    {
        "story_id": "US-007",
        "title": "Patient Appointment Booking",
        "domain": "Healthcare",
        "priority": "High",
        "story_points": 13,
        "user_story": "As a patient, I want to book, reschedule, or cancel an appointment with a doctor online, so that I don't have to call the clinic during business hours.",
        "acceptance_criteria": (
            "1. Patient sees real-time availability slots for selected doctor, date, and location.\n"
            "2. Double-booking is prevented — slot locks for 5 minutes once selected.\n"
            "3. Booking confirmation sent via email and SMS.\n"
            "4. Reschedule triggers cancellation of old slot and booking of new slot atomically.\n"
            "5. Cancellation up to 24 hours before appointment is free; < 24 hours incurs $25 fee.\n"
            "6. Appointment history viewable with filters by date range and status."
        ),
        "frs": (
            "FR-043: The system shall expose GET /doctors/{id}/slots?date=YYYY-MM-DD returning "
            "available time windows with status (available/locked/booked).\n"
            "FR-044: When a patient selects a slot, the system shall set a pessimistic lock "
            "(status=locked, locked_until=now+5min) to prevent double-booking.\n"
            "FR-045: The booking endpoint POST /appointments shall accept doctor_id, slot_id, "
            "patient_id, and reason; it shall verify the slot is still locked by the same patient "
            "before confirming.\n"
            "FR-046: The reschedule endpoint PUT /appointments/{id} shall cancel the old slot "
            "and book the new slot within a single database transaction.\n"
            "FR-047: Cancellation endpoint DELETE /appointments/{id} shall check the 24-hour rule; "
            "if within 24 hours, a $25 fee record is created on the patient's billing account.\n"
            "FR-048: The system shall send transactional email and SMS via the notification service "
            "for booking confirmation, reschedule confirmation, and cancellation.\n"
            "FR-049: The appointment history endpoint GET /appointments?start=...&end=...&status=... "
            "shall return paginated results with doctor name, date, time, status, and location."
        ),
        "test_cases": [
            {
                "tc_id": "TC-007-01",
                "description": "Verify booking an available slot.",
                "test_type": "Functional",
                "steps": "1. Select doctor, date. 2. Choose slot 10:00–10:30. 3. Enter reason. 4. Confirm booking.",
                "expected": "Slot status → 'booked'. Confirmation email and SMS received. Appointment appears in patient's history.",
            },
            {
                "tc_id": "TC-007-02",
                "description": "Verify slot lock prevents double-booking.",
                "test_type": "Functional",
                "steps": "1. Patient A selects 10:00 slot (lock acquired). 2. Patient B queries same slot within 5 minutes.",
                "expected": "Patient B sees slot as 'unavailable' (locked). Cannot select it.",
            },
            {
                "tc_id": "TC-007-03",
                "description": "Verify slot lock expires and slot becomes available again.",
                "test_type": "Functional",
                "steps": "1. Patient A selects slot but does not confirm booking. 2. Wait 6 minutes. 3. Patient B queries slots.",
                "expected": "Slot status reverts to 'available'. Patient B can book it.",
            },
            {
                "tc_id": "TC-007-04",
                "description": "Verify cancellation with late fee.",
                "test_type": "Functional",
                "steps": "1. Book appointment for tomorrow at 9:00 AM (less than 24h away). 2. Cancel it.",
                "expected": "Cancellation confirmed. $25 fee added to patient billing account. Notification sent.",
            },
            {
                "tc_id": "TC-007-05",
                "description": "Verify atomic reschedule.",
                "test_type": "Functional",
                "steps": "1. Reschedule existing appointment to new slot. 2. Simulate new slot being taken between validation and commit.",
                "expected": "Transaction rolls back. Old appointment remains intact. Error: 'Selected slot is no longer available.'",
            },
        ],
        "dependencies": "Doctor availability service; notification service (email + SMS); billing service",
        "sprint": "Sprint 3",
        "status": "In Progress",
    },
    # 8
    {
        "story_id": "US-008",
        "title": "HIPAA-Compliant Patient Data Export",
        "domain": "Healthcare",
        "priority": "High",
        "story_points": 8,
        "user_story": "As a patient, I want to export my medical records in a portable format (PDF/CSV), so that I can share them with another provider or keep them for my personal records.",
        "acceptance_criteria": (
            "1. Patient can request export from the 'My Records' page.\n"
            "2. System generates a ZIP containing PDF (human-readable) and FHIR JSON (machine-readable).\n"
            "3. Download link is sent via email; link expires after 24 hours and is single-use.\n"
            "4. Export event is logged in the HIPAA audit trail.\n"
            "5. Request is processed asynchronously; status is shown (queued → processing → ready).\n"
            "6. Patient must re-authenticate before download."
        ),
        "frs": (
            "FR-050: The system shall provide POST /export-requests that creates an export job "
            "and enqueues it to a background worker queue.\n"
            "FR-051: The export worker shall aggregate data from multiple services (allergies, "
            "medications, lab results, visit summaries) and generate a FHIR-compliant JSON bundle.\n"
            "FR-052: The worker shall also render a human-readable PDF summary using a template engine.\n"
            "FR-053: Generated files shall be zipped and uploaded to encrypted object storage "
            "(S3 with SSE-KMS) with a presigned URL valid for 24 hours and limited to 1 access.\n"
            "FR-054: The system shall send the download link via the notification service and "
            "record the export in a HIPAA-compliant audit table (patient_id, timestamp, "
            "export_type, requester_ip, success/failure).\n"
            "FR-055: The download endpoint GET /exports/{id}/download shall require re-authentication "
            "(password or biometric) and verify the presigned URL has not expired.\n"
            "FR-056: The system shall enforce that only the requesting patient (or a delegated "
            "guardian) can access the export."
        ),
        "test_cases": [
            {
                "tc_id": "TC-008-01",
                "description": "Verify successful export generation and download.",
                "test_type": "Functional",
                "steps": "1. Navigate to My Records → Request Export. 2. Wait for 'Ready' status. 3. Click download link in email. 4. Re-authenticate. 5. Open ZIP.",
                "expected": "ZIP contains PDF with visit summaries/labs and FHIR JSON. Audit log entry created.",
            },
            {
                "tc_id": "TC-008-02",
                "description": "Verify download link expiration.",
                "test_type": "Functional",
                "steps": "1. Request export. 2. Wait 25 hours. 3. Click download link.",
                "expected": "Error: 'This download link has expired. Please request a new export.' Audit log records attempted access.",
            },
            {
                "tc_id": "TC-008-03",
                "description": "Verify download link is single-use.",
                "test_type": "Functional",
                "steps": "1. Request export. 2. Download successfully. 3. Attempt to download again with same link.",
                "expected": "Error: 'This download link has already been used.'",
            },
            {
                "tc_id": "TC-008-04",
                "description": "Verify cross-patient access is blocked.",
                "test_type": "Security",
                "steps": "1. Patient A requests export. 2. Patient B attempts GET /exports/{patientA_export_id}/download.",
                "expected": "HTTP 403. Audit log records unauthorized access attempt.",
            },
            {
                "tc_id": "TC-008-05",
                "description": "Verify re-authentication gate.",
                "test_type": "Security",
                "steps": "1. Click download link while authenticated with a stale session (30+ min). 2. Attempt download.",
                "expected": "Redirect to re-authentication screen. Only after valid credentials does download proceed.",
            },
        ],
        "dependencies": "FHIR data aggregation service; PDF rendering service; S3 with SSE-KMS; notification service",
        "sprint": "Sprint 5",
        "status": "Backlog",
    },
    # 9
    {
        "story_id": "US-009",
        "title": "Social Media Feed with Algorithmic Ranking",
        "domain": "Social Media",
        "priority": "High",
        "story_points": 21,
        "user_story": "As a user, I want a personalized feed that shows posts ranked by relevance (engagement, recency, my interests), so that I see the most interesting content first.",
        "acceptance_criteria": (
            "1. Feed loads the top 20 ranked posts in ≤ 1 second.\n"
            "2. Ranking algorithm factors: recency (30%), engagement signals (40%), user affinity (30%).\n"
            "3. User can toggle between 'Algorithmic' and 'Chronological' feed views.\n"
            "4. Infinite scroll loads the next page automatically.\n"
            "5. Feed excludes posts from muted/blocked users.\n"
            "6. New posts badge appears; tapping it refreshes the feed without losing scroll position."
        ),
        "frs": (
            "FR-057: The system shall implement a feed-ranking service that scores each candidate post "
            "using a weighted formula: score = 0.3 × recency_score + 0.4 × engagement_score + "
            "0.3 × affinity_score.\n"
            "FR-058: Recency score shall decay exponentially with post age: e^(-λ × age_hours), "
            "where λ is configurable.\n"
            "FR-059: Engagement score shall be a normalized function of likes, comments, shares, "
            "and view duration in the past 24 hours.\n"
            "FR-060: Affinity score shall be derived from a user-user collaborative filtering model "
            "or interest-tag overlap with the post author/topics.\n"
            "FR-061: The GET /feed endpoint shall accept a ranking_mode param (algorithmic|chronological) "
            "and a cursor for pagination; default is algorithmic.\n"
            "FR-062: The system shall pre-compute and cache personalized feed top-N in Redis for each "
            "active user every 5 minutes.\n"
            "FR-063: The system shall filter out posts from muted/blocked relationships at query time.\n"
            "FR-064: The client shall implement Intersection Observer-based infinite scroll, requesting "
            "the next cursor page when the sentinel element is within the viewport."
        ),
        "test_cases": [
            {
                "tc_id": "TC-009-01",
                "description": "Verify algorithmic feed loads with correct ranking.",
                "test_type": "Functional",
                "steps": "1. Seed test data: posts with varying ages, engagement, and affinity. 2. Load feed for User A.",
                "expected": "Posts ranked in descending order of computed score. High-affinity, high-engagement recent post appears first.",
            },
            {
                "tc_id": "TC-009-02",
                "description": "Verify chronological feed toggle.",
                "test_type": "Functional",
                "steps": "1. Load algorithmic feed. 2. Toggle to 'Chronological'. 3. Observe post order.",
                "expected": "Post order is strictly reverse-chronological (newest first). Toggle persists across page refreshes.",
            },
            {
                "tc_id": "TC-009-03",
                "description": "Verify muted user's posts are excluded.",
                "test_type": "Functional",
                "steps": "1. User A mutes User B. 2. User B makes a post. 3. User A loads feed.",
                "expected": "User B's post does not appear in User A's feed at any scroll depth.",
            },
            {
                "tc_id": "TC-009-04",
                "description": "Verify infinite scroll loads next page.",
                "test_type": "Functional",
                "steps": "1. Load feed. 2. Scroll to bottom. 3. Observe loading spinner → new posts.",
                "expected": "Next 20 posts appended to feed. No duplicate posts. URL/fragment updated for deep linking.",
            },
            {
                "tc_id": "TC-009-05",
                "description": "Verify feed load time SLA.",
                "test_type": "Performance",
                "steps": "1. Cache miss scenario: cold start for User X with 500k candidate posts. 2. Measure time to first byte.",
                "expected": "p95 TTFB ≤ 1000 ms. Redis cache hit ratio ≥ 95% after warm-up.",
            },
        ],
        "dependencies": "Post service; user graph service; Redis caching layer; ML model serving for affinity scores",
        "sprint": "Sprint 7",
        "status": "Backlog",
    },
    # 10
    {
        "story_id": "US-010",
        "title": "Employee Leave Management Workflow",
        "domain": "HR & Payroll",
        "priority": "Medium",
        "story_points": 8,
        "user_story": "As an employee, I want to apply for leave, track approval status, and view my leave balance, so that I can plan my time off without manual coordination with HR.",
        "acceptance_criteria": (
            "1. Employee selects leave type (sick, casual, earned), dates, and optional reason.\n"
            "2. System validates against leave balance; warns if insufficient.\n"
            "3. Request routes to the employee's reporting manager for approval.\n"
            "4. Manager receives email/push notification and can approve/reject with comments.\n"
            "5. Employee notified of decision; leave balance updated on approval.\n"
            "6. Leave calendar view shows team-wide absences for the month."
        ),
        "frs": (
            "FR-065: The system shall provide POST /leave-requests with body: leave_type, start_date, "
            "end_date, reason, and optional attachment.\n"
            "FR-066: The server shall validate the date range (start ≤ end, no overlap with existing "
            "approved leave, dates not in the past) and check available leave balance from the leave "
            "entitlement service.\n"
            "FR-067: After validation, the system shall create a leave request with status 'pending' "
            "and push a notification to the reporting manager via the notification hub.\n"
            "FR-068: The manager approval endpoint PUT /leave-requests/{id}/decision shall accept "
            "status=approved|rejected and comments; on approval, call the leave-balance service "
            "to decrement entitlement.\n"
            "FR-069: The system shall send the decision notification to the employee's registered "
            "email and device.\n"
            "FR-070: The team calendar endpoint GET /leave-calendar?month=YYYY-MM shall return all "
            "approved leave entries for the requesting user's department, aggregated by date.\n"
            "FR-071: The system shall escalate unactioned requests to the skip-level manager after "
            "48 hours via a scheduled job."
        ),
        "test_cases": [
            {
                "tc_id": "TC-010-01",
                "description": "Verify successful leave application and approval flow.",
                "test_type": "Functional",
                "steps": "1. Employee applies for 3 days casual leave (balance: 5). 2. Manager receives notification. 3. Manager approves with comment 'Enjoy!'. 4. Employee checks status and balance.",
                "expected": "Status → 'Approved'. Balance → 2. Employee receives approval notification.",
            },
            {
                "tc_id": "TC-010-02",
                "description": "Verify rejection flow with comments.",
                "test_type": "Functional",
                "steps": "1. Employee applies for leave. 2. Manager rejects with 'Critical sprint — please reschedule.'",
                "expected": "Status → 'Rejected'. Employee notified with rejection reason. Balance unchanged.",
            },
            {
                "tc_id": "TC-010-03",
                "description": "Verify insufficient balance warning.",
                "test_type": "Functional",
                "steps": "1. Employee has 1 day sick leave balance. 2. Apply for 3 days sick leave.",
                "expected": "Validation warning: 'You have only 1 day(s) remaining. Request will be flagged.' Request still submittable but flagged.",
            },
            {
                "tc_id": "TC-010-04",
                "description": "Verify overlapping leave prevention.",
                "test_type": "Functional",
                "steps": "1. Employee has approved leave Dec 10–12. 2. Attempt to apply for Dec 11–14.",
                "expected": "Error: 'Your leave overlaps with an existing approved leave (Dec 10–12).'",
            },
            {
                "tc_id": "TC-010-05",
                "description": "Verify escalation after 48 hours.",
                "test_type": "Functional",
                "steps": "1. Apply for leave. 2. Advance system clock 49 hours. 3. Trigger escalation job.",
                "expected": "Skip-level manager receives escalation notification. Original manager marked as 'missed SLA'.",
            },
            {
                "tc_id": "TC-010-06",
                "description": "Verify team calendar shows department absences.",
                "test_type": "Functional",
                "steps": "1. Two employees in same dept have overlapping approved leave. 2. Open team calendar for that month.",
                "expected": "Both employees appear on overlapping dates. Calendar renders correctly.",
            },
        ],
        "dependencies": "Leave balance service; organizational hierarchy service; notification hub; scheduled job framework",
        "sprint": "Sprint 4",
        "status": "Done",
    },
    # 11
    {
        "story_id": "US-011",
        "title": "Real-Time Dashboard with WebSocket Data",
        "domain": "Analytics / BI",
        "priority": "Medium",
        "story_points": 13,
        "user_story": "As a business analyst, I want a real-time dashboard that shows live KPIs (revenue, active users, order volume) updated via WebSocket, so that I can monitor business performance without refreshing the page.",
        "acceptance_criteria": (
            "1. Dashboard displays 6 KPI cards: Revenue, Orders, Active Users, Conversion Rate, Avg. Order Value, Cart Abandonment.\n"
            "2. KPI values update in real-time via WebSocket connection; fallback to polling if WebSocket unavailable.\n"
            "3. Each KPI shows a sparkline chart (last 24 hours) and % change vs. previous period.\n"
            "4. User can select time range: Today, This Week, This Month, Custom.\n"
            "5. Dashboard auto-reconnects WebSocket on disconnect with exponential backoff.\n"
            "6. Data freshness indicator shows 'Live' (green), 'Delayed' (amber > 30s), 'Stale' (red > 5 min)."
        ),
        "frs": (
            "FR-072: The system shall establish a WebSocket connection at wss://<host>/ws/dashboard "
            "authenticated via a JWT token passed in the connection query string.\n"
            "FR-073: The server shall push aggregated KPI updates from a Kafka/event-stream topic "
            "to connected dashboard clients at a maximum frequency of once per second per client.\n"
            "FR-074: The server shall implement a fallback HTTP polling endpoint "
            "GET /dashboard/kpis?range=today|week|month which returns the same KPI payload.\n"
            "FR-075: The client shall detect WebSocket connection drops and reconnect using "
            "exponential backoff: initial delay 1s, max 30s, with jitter.\n"
            "FR-076: KPI computation shall be performed by a stream processor (e.g., Kafka Streams "
            "or Flink) that aggregates raw events into rolling-window metrics.\n"
            "FR-077: Sparkline data shall be fetched via GET /dashboard/sparklines?metric={name}&range={range} "
            "returning an array of {timestamp, value} pairs for the last 24 hours at 5-minute intervals.\n"
            "FR-078: The data freshness monitor shall track the timestamp of the last successfully "
            "processed event and compare against the current wall-clock time to determine the indicator state."
        ),
        "test_cases": [
            {
                "tc_id": "TC-011-01",
                "description": "Verify KPI cards update in real-time.",
                "test_type": "Functional",
                "steps": "1. Open dashboard. 2. Place a new order via API. 3. Observe Orders KPI.",
                "expected": "Orders KPI increments within 2 seconds of order placement. Revenue and Active Users update concurrently.",
            },
            {
                "tc_id": "TC-011-02",
                "description": "Verify WebSocket auto-reconnect.",
                "test_type": "Functional",
                "steps": "1. Open dashboard with active WebSocket. 2. Kill the WebSocket server. 3. Wait 5 seconds. 4. Restart server.",
                "expected": "Connection status changes to 'Disconnected'. Client retries with backoff. Reconnects when server is back. No data gap.",
            },
            {
                "tc_id": "TC-011-03",
                "description": "Verify fallback to HTTP polling.",
                "test_type": "Functional",
                "steps": "1. Block WebSocket port at network level. 2. Open dashboard.",
                "expected": "WebSocket connection fails. Client falls back to HTTP polling every 5s. KPIs still display and update.",
            },
            {
                "tc_id": "TC-011-04",
                "description": "Verify sparkline chart data.",
                "test_type": "Functional",
                "steps": "1. Load dashboard. 2. Inspect sparkline chart for Revenue.",
                "expected": "Chart renders 288 data points (24h / 5min intervals). Hovering shows tooltip with timestamp and value.",
            },
            {
                "tc_id": "TC-011-05",
                "description": "Verify freshness indicator states.",
                "test_type": "Functional",
                "steps": "1. Load dashboard → indicator green. 2. Stop event pipeline for 35s → amber. 3. Stop for 6 min → red.",
                "expected": "Indicator transitions green → amber at 30s → red at 5min. Tooltip shows 'Last updated: X seconds ago'.",
            },
        ],
        "dependencies": "Kafka event stream; stream processor (Flink/Kafka Streams); WebSocket gateway; JWT auth service",
        "sprint": "Sprint 6",
        "status": "In Progress",
    },
    # 12
    {
        "story_id": "US-012",
        "title": "Multi-Factor Authentication Enrollment",
        "domain": "Cybersecurity",
        "priority": "Critical",
        "story_points": 8,
        "user_story": "As a user, I want to enable multi-factor authentication (TOTP app or security key) on my account, so that my account is protected against unauthorized access even if my password is compromised.",
        "acceptance_criteria": (
            "1. User can enroll TOTP (Google Authenticator / Authy) via QR code scan.\n"
            "2. User can enroll a FIDO2/WebAuthn security key.\n"
            "3. After enrollment, MFA is required on next login.\n"
            "4. User is provided 10 one-time recovery codes for backup access.\n"
            "5. Enrollment requires current password re-verification.\n"
            "6. User can view and revoke enrolled MFA methods from Security Settings."
        ),
        "frs": (
            "FR-079: The system shall generate a TOTP secret (RFC 6238) per user, store it AES-256-GCM "
            "encrypted in the database, and present it as a QR code (otpauth:// URI) for app scanning.\n"
            "FR-080: The system shall support FIDO2/WebAuthn registration by generating a challenge, "
            "receiving the attestation response from the client, and storing the credential ID and "
            "public key.\n"
            "FR-081: During enrollment, the system shall first require the user's current password "
            "verification before proceeding to MFA method setup.\n"
            "FR-082: Upon successful enrollment of the first MFA method, the system shall generate "
            "10 cryptographically random recovery codes (8 alphanumeric chars each), hash them with "
            "SHA-256, and store the hashes.\n"
            "FR-083: The recovery codes shall be displayed once to the user with a 'Download as TXT' "
            "option and a warning that they will not be shown again.\n"
            "FR-084: The MFA management endpoint GET /user/security/mfa shall list enrolled methods "
            "with type, label, and creation date; DELETE /user/security/mfa/{method_id} shall allow "
            "removal after password re-verification.\n"
            "FR-085: After enrollment, the next login shall require the user to provide an MFA challenge "
            "response in addition to password."
        ),
        "test_cases": [
            {
                "tc_id": "TC-012-01",
                "description": "Verify TOTP enrollment and login flow.",
                "test_type": "Functional",
                "steps": "1. Log in. 2. Go to Security Settings → Enable MFA. 3. Re-enter password. 4. Scan QR code with authenticator app. 5. Enter TOTP code to confirm. 6. Log out and log in again.",
                "expected": "After enrollment: recovery codes displayed. Next login: prompted for TOTP after password. Login succeeds with valid TOTP.",
            },
            {
                "tc_id": "TC-012-02",
                "description": "Verify recovery code bypass.",
                "test_type": "Functional",
                "steps": "1. Enroll TOTP. 2. Log out. 3. At login, click 'Use Recovery Code'. 4. Enter a valid recovery code.",
                "expected": "Login succeeds. Used recovery code is invalidated (cannot be reused).",
            },
            {
                "tc_id": "TC-012-03",
                "description": "Verify invalid TOTP code is rejected.",
                "test_type": "Functional",
                "steps": "1. Enroll TOTP. 2. At login, enter incorrect 6-digit code 3 times.",
                "expected": "First 3 attempts show 'Invalid code. X attempts remaining.' 4th attempt locks account for 15 minutes.",
            },
            {
                "tc_id": "TC-012-04",
                "description": "Verify MFA method revocation.",
                "test_type": "Functional",
                "steps": "1. Enroll TOTP. 2. Go to Security Settings. 3. Remove TOTP method. 4. Log out and log in.",
                "expected": "TOTP removed. Login only requires password (no MFA challenge). Recovery codes invalidated.",
            },
            {
                "tc_id": "TC-012-05",
                "description": "Verify password re-verification is required for enrollment.",
                "test_type": "Security",
                "steps": "1. Log in. 2. Navigate to Enable MFA page. 3. Enter wrong password before MFA setup.",
                "expected": "Error: 'Password is incorrect.' MFA setup is blocked. Audit log records attempt.",
            },
        ],
        "dependencies": "Auth service; WebAuthn/FIDO2 library; QR code generation library",
        "sprint": "Sprint 2",
        "status": "Done",
    },
    # 13
    {
        "story_id": "US-013",
        "title": "Bulk CSV Import for Inventory Management",
        "domain": "Supply Chain",
        "priority": "Medium",
        "story_points": 8,
        "user_story": "As a warehouse manager, I want to import inventory data via CSV upload, so that I can update thousands of SKUs at once instead of manually entering each one.",
        "acceptance_criteria": (
            "1. User uploads a CSV file with columns: SKU, Name, Quantity, UnitPrice, Warehouse.\n"
            "2. System validates CSV headers, data types, and business rules (SKU uniqueness, quantity ≥ 0).\n"
            "3. Valid rows are processed; invalid rows are returned in an error report CSV.\n"
            "4. Import progress is shown via a progress bar polling a job status endpoint.\n"
            "5. File size limit: 10 MB (~100k rows).\n"
            "6. Import history shows past imports with date, file name, success/error counts."
        ),
        "frs": (
            "FR-086: The system shall provide POST /inventory/import that accepts a multipart/form-data "
            "upload with a CSV file and returns a job_id for async processing.\n"
            "FR-087: The server shall parse the CSV in a background worker; validate header rows to "
            "match required columns; reject the entire import if headers are missing or malformed.\n"
            "FR-088: For each row, apply validation rules: SKU is non-empty and unique in the system "
            "(or within the batch for new SKUs); Quantity is a non-negative integer; UnitPrice is a "
            "positive decimal ≤ 999,999.99; Warehouse must match a known warehouse code.\n"
            "FR-089: Valid rows shall be upserted (INSERT ... ON CONFLICT UPDATE) into the inventory "
            "table within a batch transaction.\n"
            "FR-090: Invalid rows shall be collected with row number and error messages; on job completion, "
            "the system generates an error CSV and stores it for download via "
            "GET /inventory/imports/{job_id}/errors.\n"
            "FR-091: Job progress shall be exposed via GET /inventory/imports/{job_id}/status returning "
            "{ status: pending|processing|completed|failed, total_rows, processed_rows, error_count }.\n"
            "FR-092: File size shall be limited to 10 MB; the system shall reject uploads exceeding "
            "this limit with HTTP 413 Payload Too Large.\n"
            "FR-093: Import history endpoint GET /inventory/imports shall return a paginated list of "
            "past imports with metadata."
        ),
        "test_cases": [
            {
                "tc_id": "TC-013-01",
                "description": "Verify successful CSV import.",
                "test_type": "Functional",
                "steps": "1. Upload CSV with 3 valid SKUs. 2. Poll job status until completed. 3. Query inventory for imported SKUs.",
                "expected": "Job completes with status 'completed'. All 3 SKUs present in inventory with correct data.",
            },
            {
                "tc_id": "TC-013-02",
                "description": "Verify partial success with error report.",
                "test_type": "Functional",
                "steps": "1. Upload CSV with 5 rows: 2 have negative quantity, 1 has invalid warehouse. 2. Wait for completion. 3. Download error CSV.",
                "expected": "2 valid rows processed. 3 rows in error CSV with row number and reason. Error count = 3.",
            },
            {
                "tc_id": "TC-013-03",
                "description": "Verify file size limit enforcement.",
                "test_type": "Boundary",
                "steps": "1. Attempt to upload a 10.5 MB CSV file.",
                "expected": "HTTP 413 returned. Upload rejected before processing begins.",
            },
            {
                "tc_id": "TC-013-04",
                "description": "Verify missing headers rejection.",
                "test_type": "Functional",
                "steps": "1. Upload CSV with headers: SKU, Name, Quantity (missing UnitPrice, Warehouse).",
                "expected": "Job fails immediately. Error: 'Missing required columns: UnitPrice, Warehouse.' No rows processed.",
            },
            {
                "tc_id": "TC-013-05",
                "description": "Verify upsert behavior — existing SKU updated.",
                "test_type": "Functional",
                "steps": "1. SKU 'ABC-123' exists with quantity 50. 2. Import CSV with SKU 'ABC-123', quantity 100.",
                "expected": "SKU 'ABC-123' quantity updated to 100. No duplicate record created.",
            },
        ],
        "dependencies": "Background job queue (e.g., Bull/BullMQ, Sidekiq); object storage for file uploads; inventory database",
        "sprint": "Sprint 7",
        "status": "Backlog",
    },
    # 14
    {
        "story_id": "US-014",
        "title": "Push Notification Preferences Management",
        "domain": "Mobile / Engagement",
        "priority": "Low",
        "story_points": 5,
        "user_story": "As a mobile app user, I want to manage which types of push notifications I receive (promotions, order updates, social activity), so that I only get alerts that matter to me.",
        "acceptance_criteria": (
            "1. Settings screen shows toggle switches for each notification category.\n"
            "2. Default for all categories is ON for new users.\n"
            "3. Changes are saved immediately and take effect on next notification dispatch.\n"
            "4. 'Order Updates' (transactional) cannot be disabled — toggle is locked.\n"
            "5. User can also set quiet hours (start time, end time, timezone) during which no notifications are delivered.\n"
            "6. Preferences sync across user's devices."
        ),
        "frs": (
            "FR-094: The system shall maintain a user notification preferences record with boolean "
            "fields for each category: promotions, order_updates, social_activity, product_recommendations, "
            "account_alerts.\n"
            "FR-095: The PATCH /user/preferences/notifications endpoint shall accept a partial update "
            "payload; only provided fields are updated; order_updates shall be rejected if set to false.\n"
            "FR-096: The notification dispatch service shall filter recipients by checking their "
            "preferences before sending; order-triggered notifications shall bypass this filter.\n"
            "FR-097: The quiet-hours configuration shall be stored as start_time, end_time (HH:MM 24h), "
            "and timezone (IANA format); notifications queued during quiet hours shall be held and "
            "delivered at the end of the quiet period.\n"
            "FR-098: On preferences update, the server shall publish a change event to a real-time "
            "channel to sync across the user's active sessions/devices within 5 seconds.\n"
            "FR-099: New users shall be provisioned with a default preferences record (all true) "
            "upon account creation."
        ),
        "test_cases": [
            {
                "tc_id": "TC-014-01",
                "description": "Verify disabling a notification category.",
                "test_type": "Functional",
                "steps": "1. Go to Notification Settings. 2. Toggle 'Promotions' OFF. 3. Trigger a promotional notification.",
                "expected": "Promotional notification is not delivered. Other categories still deliver normally.",
            },
            {
                "tc_id": "TC-014-02",
                "description": "Verify order update toggle is locked ON.",
                "test_type": "Functional",
                "steps": "1. Go to Notification Settings. 2. Observe 'Order Updates' toggle.",
                "expected": "Toggle is ON and disabled (grayed out). Tooltip: 'Transactional notifications cannot be disabled.'",
            },
            {
                "tc_id": "TC-014-03",
                "description": "Verify API blocks disabling order_updates.",
                "test_type": "Security",
                "steps": "1. Send PATCH /user/preferences/notifications with {\"order_updates\": false}.",
                "expected": "HTTP 422. Error: 'order_updates cannot be disabled.' Preference remains true.",
            },
            {
                "tc_id": "TC-014-04",
                "description": "Verify quiet hours delay delivery.",
                "test_type": "Functional",
                "steps": "1. Set quiet hours 22:00–07:00. 2. Trigger promotional notification at 23:00. 3. Check device at 23:05.",
                "expected": "No notification delivered during quiet hours. Notification arrives at 07:00.",
            },
            {
                "tc_id": "TC-014-05",
                "description": "Verify cross-device sync.",
                "test_type": "Functional",
                "steps": "1. Log in on Device A and Device B. 2. On Device A, toggle 'Social Activity' OFF. 3. Check Device B settings screen.",
                "expected": "Device B reflects updated preference within 5 seconds via real-time sync.",
            },
        ],
        "dependencies": "Push notification service (FCM/APNs); real-time sync channel (WebSocket/SSE)",
        "sprint": "Sprint 4",
        "status": "Done",
    },
    # 15
    {
        "story_id": "US-015",
        "title": "Automated Invoice Generation and Delivery",
        "domain": "Finance / Billing",
        "priority": "High",
        "story_points": 13,
        "user_story": "As a finance team member, I want the system to automatically generate and email invoices on a recurring schedule, so that we reduce manual billing effort and ensure timely delivery.",
        "acceptance_criteria": (
            "1. Invoices are generated on the 1st of each month for all active subscriptions.\n"
            "2. Invoice PDF includes: company logo, invoice number, billing period, line items, subtotal, tax, total, payment due date.\n"
            "3. Invoice number format: INV-YYYYMM-XXXXX (sequential).\n"
            "4. Email is sent with PDF attachment and a payment link.\n"
            "5. Failed generation is retried up to 3 times with exponential backoff; after 3 failures, ops team is alerted.\n"
            "6. Finance team can trigger ad-hoc invoice generation for a specific customer."
        ),
        "frs": (
            "FR-100: The system shall have a scheduled job (cron: 0 2 1 * *) that queries all active "
            "subscriptions and enqueues invoice generation tasks for each.\n"
            "FR-101: Each invoice generation task shall compute line items from the subscription's "
            "usage records for the previous billing period, apply applicable taxes (tax rate from "
            "customer's jurisdiction), and calculate total.\n"
            "FR-102: Invoice numbers shall be generated atomically using a database sequence; "
            "format: INV-{YYYYMM}-{zero-padded 5-digit sequence}.\n"
            "FR-103: The system shall render the invoice as a PDF using a template engine "
            "(e.g., Gotenberg, Puppeteer, or Prawn) with the company branding and all required fields.\n"
            "FR-104: The PDF shall be stored in durable object storage and linked to the invoice "
            "record in the database.\n"
            "FR-105: The system shall send the invoice email via the notification service with the "
            "PDF attachment (base64-encoded or via a presigned download link valid 30 days).\n"
            "FR-106: Failed invoice generation shall be retried by the background job framework with "
            "exponential backoff (1m, 5m, 15m). After the 3rd failure, the system shall create an "
            "incident ticket via PagerDuty/Opsgenie webhook.\n"
            "FR-107: Ad-hoc generation shall be triggered via POST /invoices/generate with a "
            "subscription_id parameter, restricted to users with 'finance:invoice' permission."
        ),
        "test_cases": [
            {
                "tc_id": "TC-015-01",
                "description": "Verify scheduled monthly invoice generation.",
                "test_type": "Functional",
                "steps": "1. Set up 3 active subscriptions. 2. Advance system clock to 1st of next month 02:00. 3. Observe job execution.",
                "expected": "3 invoices generated with sequential invoice numbers. Each PDF correctly reflects usage. 3 emails sent.",
            },
            {
                "tc_id": "TC-015-02",
                "description": "Verify invoice number uniqueness and format.",
                "test_type": "Functional",
                "steps": "1. Generate 10 invoices in the same month. 2. Inspect invoice numbers.",
                "expected": "Format INV-202607-00001 through INV-202607-00010. No gaps or duplicates.",
            },
            {
                "tc_id": "TC-015-03",
                "description": "Verify PDF content accuracy.",
                "test_type": "Functional",
                "steps": "1. Generate invoice for known subscription with $100 base + $20 usage. 2. Open generated PDF.",
                "expected": "PDF shows: line items ($100 + $20), tax (e.g., 8% → $9.60), total ($129.60), due date (net-30).",
            },
            {
                "tc_id": "TC-015-04",
                "description": "Verify retry and alerting on failure.",
                "test_type": "Functional",
                "steps": "1. Configure PDF service to fail (return 500). 2. Trigger invoice generation. 3. Observe retry behavior and alert after 3rd failure.",
                "expected": "Retries at 1m, 5m, 15m. After 3rd failure, PagerDuty alert fired. Invoice status → 'failed'.",
            },
            {
                "tc_id": "TC-015-05",
                "description": "Verify ad-hoc invoice generation by authorized user.",
                "test_type": "Functional",
                "steps": "1. Log in as finance team member. 2. Send POST /invoices/generate with valid subscription_id. 3. Check invoice.",
                "expected": "Invoice generated and sent immediately. Appears in invoice list with source 'manual'.",
            },
            {
                "tc_id": "TC-015-06",
                "description": "Verify unauthorized ad-hoc generation is blocked.",
                "test_type": "Security",
                "steps": "1. Log in as regular user without finance permission. 2. Send POST /invoices/generate.",
                "expected": "HTTP 403. No invoice created.",
            },
        ],
        "dependencies": "PDF rendering service; subscription billing database; tax rate service; notification service; PagerDuty/Opsgenie integration",
        "sprint": "Sprint 8",
        "status": "Backlog",
    },
    # 16
    {
        "story_id": "US-016",
        "title": "Rate Limiting and API Abuse Protection",
        "domain": "Platform / Infrastructure",
        "priority": "High",
        "story_points": 8,
        "user_story": "As a platform engineer, I want to implement rate limiting on public APIs to prevent abuse and ensure fair usage, so that the service remains available for all legitimate users.",
        "acceptance_criteria": (
            "1. Rate limits are enforced per API key / user token.\n"
            "2. Default limit: 100 requests per minute per user; configurable per endpoint and tier.\n"
            "3. When limit exceeded, HTTP 429 returned with Retry-After header.\n"
            "4. Remaining limit and reset time are returned in X-RateLimit-* response headers.\n"
            "5. Rate limit counters are stored in Redis with a sliding-window algorithm.\n"
            "6. Rate limit events exceeding threshold are logged and optionally trigger alerts."
        ),
        "frs": (
            "FR-108: The system shall implement a rate-limiting middleware using the sliding-window "
            "algorithm with Redis sorted sets, keyed by user_id:endpoint.\n"
            "FR-109: Default rate limit shall be 100 requests per rolling 60-second window per user; "
            "limits shall be overridable per endpoint via a configuration table or feature-flag service.\n"
            "FR-110: When a request exceeds the limit, the middleware shall respond with "
            "HTTP 429 Too Many Requests and a Retry-After header indicating seconds until reset.\n"
            "FR-111: Every API response shall include headers: X-RateLimit-Limit, X-RateLimit-Remaining, "
            "X-RateLimit-Reset (Unix timestamp).\n"
            "FR-112: Rate limit consumption shall be atomic using Redis MULTI/EXEC or Lua scripts to "
            "prevent race conditions in a distributed environment.\n"
            "FR-113: When a user exceeds 80% of their limit, a warning event shall be emitted; "
            "at 100%, an abuse event shall be logged to the security event pipeline.\n"
            "FR-114: Administrative endpoints (GET /admin/rate-limits) shall allow viewing and "
            "temporarily adjusting limits for specific users or IPs."
        ),
        "test_cases": [
            {
                "tc_id": "TC-016-01",
                "description": "Verify rate limit enforced at 100 RPM.",
                "test_type": "Functional",
                "steps": "1. Send 100 requests within 60 seconds. 2. Send 101st request.",
                "expected": "First 100 succeed. 101st returns HTTP 429 with Retry-After header. X-RateLimit-Remaining = 0.",
            },
            {
                "tc_id": "TC-016-02",
                "description": "Verify sliding-window accuracy.",
                "test_type": "Functional",
                "steps": "1. Send 50 requests at t=0. 2. Send 50 requests at t=30s. 3. Send 1 request at t=61s (window slides).",
                "expected": "Request at t=61s succeeds (first 50 have aged out). X-RateLimit-Remaining > 0.",
            },
            {
                "tc_id": "TC-016-03",
                "description": "Verify per-endpoint limits.",
                "test_type": "Functional",
                "steps": "1. Configure /api/search limit = 30 RPM. 2. Send 30 requests to /api/search. 3. Send 31st.",
                "expected": "31st request to /api/search returns 429. Requests to other endpoints (default 100 RPM) still succeed.",
            },
            {
                "tc_id": "TC-016-04",
                "description": "Verify rate limit isolation between users.",
                "test_type": "Functional",
                "steps": "1. User A exhausts their limit. 2. User B (different API key) sends requests.",
                "expected": "User B's requests succeed — rate limits are per-user, not global.",
            },
            {
                "tc_id": "TC-016-05",
                "description": "Verify rate limit headers are present.",
                "test_type": "Functional",
                "steps": "1. Send any API request. 2. Inspect response headers.",
                "expected": "Headers present: X-RateLimit-Limit, X-RateLimit-Remaining (integer), X-RateLimit-Reset (epoch seconds).",
            },
        ],
        "dependencies": "Redis cluster; API gateway / middleware framework; configuration service",
        "sprint": "Sprint 3",
        "status": "Done",
    },
    # 17
    {
        "story_id": "US-017",
        "title": "Data Export with GDPR Right to Portability",
        "domain": "Data Privacy / Compliance",
        "priority": "High",
        "story_points": 8,
        "user_story": "As a user, I want to request a complete export of all my personal data in a machine-readable format, so that I can exercise my GDPR right to data portability.",
        "acceptance_criteria": (
            "1. User submits data export request from Privacy Settings page.\n"
            "2. Request confirmed via email to prevent unauthorized exports.\n"
            "3. System compiles all personal data: profile, orders, messages, reviews, activity logs.\n"
            "4. Data is delivered as a JSON file (structured, machine-readable).\n"
            "5. Export is available for download for 7 days after completion.\n"
            "6. SLA: export completed within 72 hours (GDPR compliance); status tracker visible.\n"
            "7. User can also request account deletion from the same page."
        ),
        "frs": (
            "FR-115: The system shall provide POST /privacy/export-request that initiates a GDPR "
            "data portability workflow for the authenticated user.\n"
            "FR-116: After receiving the request, the system shall send a confirmation email with a "
            "unique verification link; the export process starts only after the user clicks the link.\n"
            "FR-117: The export worker shall query all data stores associated with the user ID: "
            "profile, orders, payment history (masked PAN), messages, product reviews, support tickets, "
            "login history, consent records.\n"
            "FR-118: Collected data shall be serialized into a structured JSON document conforming to "
            "a defined schema (each entity as a top-level key with an array of records).\n"
            "FR-119: The JSON file shall be compressed (gzip) and stored in encrypted object storage "
            "with a presigned download URL expiring after 7 days.\n"
            "FR-120: Export job status shall be queryable via GET /privacy/export-requests returning "
            "{ status: pending_confirmation|in_progress|completed|failed, progress_percent, "
            "created_at, completed_at, download_url_expires_at }.\n"
            "FR-121: The completed notification email shall include the download link and the expiration date.\n"
            "FR-122: The account deletion endpoint POST /privacy/delete-account shall initiate a "
            "workflow that soft-deletes the account (retains data for 30 days in a 'pending_deletion' "
            "state for reversal), then permanently anonymizes or deletes all PII."
        ),
        "test_cases": [
            {
                "tc_id": "TC-017-01",
                "description": "Verify full GDPR export workflow.",
                "test_type": "Functional",
                "steps": "1. Go to Privacy Settings → Request Data Export. 2. Open confirmation email → click verify. 3. Wait for export completion. 4. Download and inspect JSON.",
                "expected": "JSON contains all data categories. Accurate and complete. Download link expires in 7 days.",
            },
            {
                "tc_id": "TC-017-02",
                "description": "Verify export requires email confirmation.",
                "test_type": "Security",
                "steps": "1. Request export. 2. Do not click email verification. 3. Check export status.",
                "expected": "Status = 'pending_confirmation'. Data not yet collected. Email not re-sent.",
            },
            {
                "tc_id": "TC-017-03",
                "description": "Verify download link expiration.",
                "test_type": "Functional",
                "steps": "1. Complete export. 2. Wait 8 days. 3. Attempt to download using original link.",
                "expected": "Error: 'This download has expired. Please submit a new request.'",
            },
            {
                "tc_id": "TC-017-04",
                "description": "Verify GDPR 72-hour SLA tracking.",
                "test_type": "Functional",
                "steps": "1. Confirm export request. 2. Observe status tracker showing elapsed time. 3. Simulate job taking > 72h.",
                "expected": "Tracker shows elapsed time. At 72h, an SLA breach alert is triggered to the DPO dashboard.",
            },
            {
                "tc_id": "TC-017-05",
                "description": "Verify account deletion with reversal window.",
                "test_type": "Functional",
                "steps": "1. Request account deletion. 2. Confirm via email link. 3. Try to log in within 30 days. 4. Recover account via support.",
                "expected": "Within 30 days: login redirects to 'Account scheduled for deletion' with reactivation option. After 30 days: PII permanently anonymized.",
            },
        ],
        "dependencies": "Email verification service; data aggregation service across microservices; encrypted object storage (S3 SSE-KMS); DPO alerting dashboard",
        "sprint": "Sprint 10",
        "status": "Backlog",
    },
    # 18
    {
        "story_id": "US-018",
        "title": "Dark Mode Toggle with Persistence",
        "domain": "UI/UX",
        "priority": "Low",
        "story_points": 3,
        "user_story": "As a user, I want to switch between light and dark theme, so that I can reduce eye strain when using the app at night.",
        "acceptance_criteria": (
            "1. Toggle switch in the user menu or settings to choose Light / Dark / System.\n"
            "2. 'System' follows the OS/browser prefers-color-scheme media query.\n"
            "3. Theme preference persists across sessions (stored in user profile and localStorage).\n"
            "4. Transition between themes is smooth (CSS transition on background/text colors).\n"
            "5. All UI components (buttons, modals, tables, forms) support both themes.\n"
            "6. Logged-out users see the theme based on localStorage or system preference."
        ),
        "frs": (
            "FR-123: The system shall support three theme modes: light, dark, and system (auto-detect).\n"
            "FR-124: The theme toggle shall be accessible from the user dropdown menu and the Settings "
            "page as a segmented control or radio group.\n"
            "FR-125: On theme change, the client shall set a data-theme attribute on the <html> element "
            "(values: 'light', 'dark') and update CSS custom properties accordingly.\n"
            "FR-126: The client shall listen for the prefers-color-scheme media query change event and "
            "switch theme dynamically when mode is 'system'.\n"
            "FR-127: For authenticated users, the theme preference shall be saved via "
            "PATCH /user/preferences { theme: 'light'|'dark'|'system' } and restored on login.\n"
            "FR-128: For unauthenticated users, the preference shall be stored in localStorage under "
            "key 'app-theme'.\n"
            "FR-129: The CSS shall define all color tokens as custom properties (e.g., --bg-primary, "
            "--text-primary) in :root and [data-theme='dark'] selectors.\n"
            "FR-130: Color transitions shall use `transition: background-color 0.3s ease, color 0.3s ease` "
            "for a smooth visual handoff."
        ),
        "test_cases": [
            {
                "tc_id": "TC-018-01",
                "description": "Verify dark mode toggle applies correctly.",
                "test_type": "Functional",
                "steps": "1. Log in. 2. Toggle theme to 'Dark'. 3. Observe all pages.",
                "expected": "Background dark (#1a1a2e or similar), text light. All components (tables, forms, modals) rendered in dark palette.",
            },
            {
                "tc_id": "TC-018-02",
                "description": "Verify system mode follows OS preference.",
                "test_type": "Functional",
                "steps": "1. Set theme to 'System'. 2. Change OS to dark mode. 3. Change OS to light mode.",
                "expected": "App theme switches to dark when OS switches, then to light. No page reload required.",
            },
            {
                "tc_id": "TC-018-03",
                "description": "Verify theme persistence across sessions.",
                "test_type": "Functional",
                "steps": "1. Set theme to 'Dark'. 2. Log out. 3. Close browser. 4. Reopen and log in.",
                "expected": "App loads in dark mode from user profile. No flicker of wrong theme (SSR hydration match).",
            },
            {
                "tc_id": "TC-018-04",
                "description": "Verify logged-out user theme from localStorage.",
                "test_type": "Functional",
                "steps": "1. As logged-out user, set theme to 'Dark'. 2. Reload the marketing page.",
                "expected": "Marketing page renders in dark mode. localStorage 'app-theme' = 'dark'.",
            },
            {
                "tc_id": "TC-018-05",
                "description": "Verify smooth transition animation.",
                "test_type": "Functional",
                "steps": "1. Toggle between light and dark mode. 2. Visually inspect transition.",
                "expected": "Background and text colors animate over ~300ms. No jarring flash.",
            },
        ],
        "dependencies": "Design system with CSS custom properties; user preferences API",
        "sprint": "Sprint 1",
        "status": "Done",
    },
    # 19
    {
        "story_id": "US-019",
        "title": "Two-Way SMS Customer Support Chat",
        "domain": "Customer Support",
        "priority": "Medium",
        "story_points": 13,
        "user_story": "As a customer, I want to communicate with support agents via SMS from my phone, so that I can get help even when I don't have the app installed or internet access.",
        "acceptance_criteria": (
            "1. Customer can text a dedicated support number and receive auto-reply with case number.\n"
            "2. SMS messages are relayed to the support agent dashboard in real-time.\n"
            "3. Agent replies from dashboard are delivered back to customer's phone as SMS.\n"
            "4. Conversation history is preserved and linked to the customer's profile.\n"
            "5. Media (images) sent via MMS are viewable by the agent.\n"
            "6. After-hours messages receive an auto-reply with business hours info and are queued for next day."
        ),
        "frs": (
            "FR-131: The system shall integrate with a telephony API (Twilio / MessageBird) via a "
            "webhook endpoint POST /webhooks/sms/inbound that receives incoming SMS messages with "
            "from_number, to_number, body, and optional media URLs.\n"
            "FR-132: On first inbound message from an unknown number, the system shall look up the "
            "customer by phone number; if not found, create a provisional contact record.\n"
            "FR-133: The system shall create (or append to) a support ticket/case and generate a "
            "unique case number, returned to the customer in an auto-reply SMS.\n"
            "FR-134: Incoming SMS messages shall be pushed to the agent dashboard in real-time via "
            "WebSocket; the agent-side chat UI shall display messages in a threaded conversation view "
            "with customer name, phone, and message bubbles.\n"
            "FR-135: Agent outbound messages shall be sent via POST to the telephony API's message "
            "resource and appended to the conversation thread.\n"
            "FR-136: Inbound MMS media URLs shall be downloaded, scanned for malware, and stored in "
            "the case attachments; the agent UI shall render images inline.\n"
            "FR-137: A scheduled job shall check business hours (configurable per location); messages "
            "received outside business hours shall receive an auto-reply template with operating hours "
            "and remain in the open queue for the next business day with an SLA clock paused."
        ),
        "test_cases": [
            {
                "tc_id": "TC-019-01",
                "description": "Verify inbound SMS creates case and auto-reply.",
                "test_type": "Functional",
                "steps": "1. From a test phone, send SMS 'My order #1234 is missing' to support number. 2. Observe phone and agent dashboard.",
                "expected": "Auto-reply SMS received: 'Thanks for contacting us! Case #CS-789 has been created...' Agent sees new conversation with the message.",
            },
            {
                "tc_id": "TC-019-02",
                "description": "Verify agent reply delivered as SMS.",
                "test_type": "Functional",
                "steps": "1. Agent types 'Let me check on that order for you' in dashboard. 2. Click Send.",
                "expected": "Customer receives SMS from support number with agent's message. Dashboard conversation thread updated.",
            },
            {
                "tc_id": "TC-019-03",
                "description": "Verify existing customer identification.",
                "test_type": "Functional",
                "steps": "1. Customer with phone +1-555-0100 (registered) sends SMS. 2. Agent views conversation.",
                "expected": "Conversation linked to existing customer profile. Agent sees customer name, email, order history in sidebar.",
            },
            {
                "tc_id": "TC-019-04",
                "description": "Verify MMS image handling.",
                "test_type": "Functional",
                "steps": "1. Send MMS with photo of damaged product to support number. 2. Agent views conversation.",
                "expected": "Agent sees image inline in chat. Image stored in case attachments. No malware detected.",
            },
            {
                "tc_id": "TC-019-05",
                "description": "Verify after-hours auto-reply.",
                "test_type": "Functional",
                "steps": "1. Send SMS at 11:00 PM (outside 8 AM–8 PM business hours). 2. Check response.",
                "expected": "Auto-reply: 'Our support team is currently offline. Business hours: 8 AM–8 PM EST. We'll respond first thing tomorrow. Case #CS-790.' SLA clock paused.",
            },
        ],
        "dependencies": "Twilio/MessageBird account and webhooks; WebSocket gateway for agent dashboard; malware scanning service; SLA tracking service",
        "sprint": "Sprint 9",
        "status": "Backlog",
    },
    # 20
    {
        "story_id": "US-020",
        "title": "Automated Regression Test Suite Trigger on PR",
        "domain": "DevOps / QA",
        "priority": "Medium",
        "story_points": 8,
        "user_story": "As a QA engineer, I want the full regression test suite to run automatically when a pull request is created, so that we catch regressions before merging code to main.",
        "acceptance_criteria": (
            "1. On PR creation or push to an open PR, CI pipeline triggers the regression suite.\n"
            "2. Test results are posted as a comment on the PR with pass/fail counts and link to full report.\n"
            "3. PR cannot be merged if any regression test fails (branch protection rule).\n"
            "4. Test execution time ≤ 30 minutes (suite parallelized across 10 shards).\n"
            "5. Flaky test detection: tests that fail non-deterministically are flagged and auto-quarantined.\n"
            "6. Test report is archived as a CI artifact for 30 days."
        ),
        "frs": (
            "FR-138: The CI/CD pipeline configuration shall define a 'regression' job that is triggered "
            "on pull_request events (opened, synchronize) targeting the main branch.\n"
            "FR-139: The test runner shall distribute tests across N parallel shards using a consistent "
            "sharding strategy (e.g., test file hash modulo N) to balance execution time.\n"
            "FR-140: Test results shall be output in JUnit XML format and published as a CI artifact; "
            "a summary comment shall be created on the PR via the SCM API (GitHub/GitLab/Bitbucket) "
            "containing total, passed, failed, skipped counts and a link to the CI run.\n"
            "FR-141: Branch protection rules shall be configured so that the 'regression' status check "
            "must pass before merge; the SCM shall enforce this blocking.\n"
            "FR-142: The system shall implement flaky test detection by tracking test outcomes across "
            "the last 10 runs; a test that passes on retry but failed on first attempt in ≥ 3 of the "
            "last 10 runs shall be flagged as flaky.\n"
            "FR-143: Flagged flaky tests shall be moved to a quarantine suite and excluded from the "
            "merge-blocking check; an issue shall be auto-created in the backlog for each quarantined test.\n"
            "FR-144: Test artifacts (reports, screenshots, logs) shall be retained in CI storage for 30 days."
        ),
        "test_cases": [
            {
                "tc_id": "TC-020-01",
                "description": "Verify regression suite triggers on PR creation.",
                "test_type": "Functional",
                "steps": "1. Create a PR against main with a code change. 2. Observe CI pipeline.",
                "expected": "'regression' job starts automatically. All 10 shards execute in parallel. PR comment posted with results.",
            },
            {
                "tc_id": "TC-020-02",
                "description": "Verify merge is blocked on test failure.",
                "test_type": "Functional",
                "steps": "1. Create PR with a change that breaks a regression test. 2. Wait for CI to complete. 3. Attempt to merge.",
                "expected": "Merge button disabled. Message: 'Required status check 'regression' is failing.'",
            },
            {
                "tc_id": "TC-020-03",
                "description": "Verify flaky test is quarantined.",
                "test_type": "Functional",
                "steps": "1. Introduce a test that passes 70% of the time. 2. Run PR CI 10 times. 3. Observe after 3 flaky failures.",
                "expected": "Test flagged in CI output as 'FLAKY'. Moved to quarantine suite. Auto-issue created in backlog. Merge no longer blocked by this test.",
            },
            {
                "tc_id": "TC-020-04",
                "description": "Verify test execution time SLA.",
                "test_type": "Performance",
                "steps": "1. Run full regression suite on a PR with 2,000 test cases across 10 shards. 2. Measure wall-clock duration.",
                "expected": "Total suite execution ≤ 30 minutes. No single shard exceeds 35 minutes (10% tolerance).",
            },
            {
                "tc_id": "TC-020-05",
                "description": "Verify test artifacts are archived.",
                "test_type": "Functional",
                "steps": "1. Run regression suite. 2. Download CI artifacts 15 days later. 3. Open JUnit XML report.",
                "expected": "Artifacts available for download. Report contains all test case details with timestamps.",
            },
        ],
        "dependencies": "CI/CD platform (GitHub Actions / Jenkins / GitLab CI); test runner framework; SCM API token; issue tracker API",
        "sprint": "Sprint 5",
        "status": "In Progress",
    },
]


# ── workbook builder ─────────────────────────────────────────────────────
def build_workbook(stories):
    wb = openpyxl.Workbook()

    # ── styles ───────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    priority_fills = {
        "Critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "High": PatternFill(start_color="FFEBC1", end_color="FFEBC1", fill_type="solid"),
        "Medium": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Low": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    }

    status_fills = {
        "Done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "In Progress": PatternFill(start_color="FFEBC1", end_color="FFEBC1", fill_type="solid"),
        "Backlog": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # ── Sheet 1: User Stories ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "User Stories"

    us_headers = [
        "Story ID", "Title", "Domain", "Priority", "Story Points",
        "Sprint", "Status", "Dependencies",
        "User Story", "Acceptance Criteria",
    ]
    for col_idx, header in enumerate(us_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, story in enumerate(stories, 2):
        values = [
            story["story_id"],
            story["title"],
            story["domain"],
            story["priority"],
            story["story_points"],
            story["sprint"],
            story["status"],
            story["dependencies"],
            story["user_story"],
            story["acceptance_criteria"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align if col_idx >= 8 else cell_align_center
            cell.border = thin_border
            if (row_idx % 2) == 0:
                if col_idx not in (4, 7):
                    cell.fill = alt_row_fill
            if col_idx == 4 and value in priority_fills:
                cell.fill = priority_fills[value]
            elif col_idx == 7 and value in status_fills:
                cell.fill = status_fills[value]

    # Column widths
    col_widths_us = [12, 40, 18, 10, 14, 10, 14, 35, 60, 60]
    for i, w in enumerate(col_widths_us, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Functional Requirements ─────────────────────────────────
    ws2 = wb.create_sheet("Functional Requirements")

    fr_headers = [
        "Story ID", "Title", "FR Reference", "Functional Requirement Specification",
    ]
    for col_idx, header in enumerate(fr_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    fr_row = 2
    for story in stories:
        fr_lines = [l.strip() for l in story["frs"].split("\n") if l.strip()]
        for line in fr_lines:
            # Extract FR-XXX reference
            fr_ref = line.split(":")[0].strip() if ":" in line else ""
            fr_desc = line.split(":", 1)[1].strip() if ":" in line else line

            row_data = [story["story_id"], story["title"], fr_ref, fr_desc]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=fr_row, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_align if col_idx == 4 else cell_align_center
                cell.border = thin_border
                if (fr_row % 2) == 0:
                    cell.fill = alt_row_fill
            fr_row += 1

    col_widths_fr = [12, 40, 14, 100]
    for i, w in enumerate(col_widths_fr, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Test Cases ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Test Cases")

    tc_headers = [
        "Story ID", "Title", "Test Case ID", "Test Description",
        "Test Type", "Test Steps", "Expected Result",
    ]
    for col_idx, header in enumerate(tc_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    tc_row = 2
    for story in stories:
        for tc in story["test_cases"]:
            values = [
                story["story_id"],
                story["title"],
                tc["tc_id"],
                tc["description"],
                tc["test_type"],
                tc["steps"],
                tc["expected"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws3.cell(row=tc_row, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_align if col_idx >= 6 else cell_align_center
                cell.border = thin_border
                if (tc_row % 2) == 0:
                    cell.fill = alt_row_fill
            tc_row += 1

    col_widths_tc = [12, 40, 14, 50, 14, 60, 60]
    for i, w in enumerate(col_widths_tc, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.auto_filter.ref = ws3.dimensions
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Summary & Metadata ──────────────────────────────────────
    ws4 = wb.create_sheet("Summary")

    ws4.merge_cells("A1:F1")
    ws4.cell(row=1, column=1, value="Dataset Summary — User Stories, FRS & Test Cases").font = Font(
        name="Calibri", size=14, bold=True, color="2F5496"
    )
    ws4.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    meta_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(meta_headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    domain_counts = {}
    priority_counts = {}
    status_counts = {}
    total_points = 0
    total_tc = 0
    total_fr = 0
    for s in stories:
        domain_counts[s["domain"]] = domain_counts.get(s["domain"], 0) + 1
        priority_counts[s["priority"]] = priority_counts.get(s["priority"], 0) + 1
        status_counts[s["status"]] = status_counts.get(s["status"], 0) + 1
        total_points += s["story_points"]
        total_tc += len(s["test_cases"])
        total_fr += len([l for l in s["frs"].split("\n") if l.strip()])

    summary_data = [
        ("Total User Stories", len(stories)),
        ("Total Functional Requirements", total_fr),
        ("Total Test Cases", total_tc),
        ("Total Story Points", total_points),
        ("Average Story Points per Story", round(total_points / len(stories), 1)),
        ("Average Test Cases per Story", round(total_tc / len(stories), 1)),
        ("Domains Covered", ", ".join(sorted(domain_counts.keys()))),
        ("", ""),
        ("Priority Distribution", ""),
        *[(f"  {k}", v) for k, v in sorted(priority_counts.items(), key=lambda x: x[1], reverse=True)],
        ("", ""),
        ("Status Distribution", ""),
        *[(f"  {k}", v) for k, v in sorted(status_counts.items(), key=lambda x: x[1], reverse=True)],
        ("", ""),
        ("Domain Breakdown", ""),
        *[(f"  {k}", v) for k, v in sorted(domain_counts.items(), key=lambda x: x[1], reverse=True)],
        ("", ""),
        ("Generated On", "2026-07-15"),
        ("Format Version", "1.0"),
    ]

    for row_idx, (metric, value) in enumerate(summary_data, 4):
        c1 = ws4.cell(row=row_idx, column=1, value=metric)
        c2 = ws4.cell(row=row_idx, column=2, value=value)
        c1.font = Font(name="Calibri", size=10, bold=("  " not in metric))
        c2.font = cell_font
        c1.alignment = cell_align
        c2.alignment = cell_align
        c1.border = thin_border
        c2.border = thin_border

    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 50

    return wb


# ── main ─────────────────────────────────────────────────────────────────
output_path = "/home/twerp/excel-prompt-dashboard/Dataset_UserStories_FRS_TestCases.xlsx"
wb = build_workbook(stories)
wb.save(output_path)
print(f"Dataset written to {output_path}")
print(f"  {len(stories)} user stories")
print(f"  {sum(len(s['test_cases']) for s in stories)} test cases")
print(f"  {sum(len([l for l in s['frs'].split(chr(10)) if l.strip()]) for s in stories)} functional requirements")
